from __future__ import annotations
"""FastAPI Client Router for multi-tenant customer-facing dashboard."""
import html
import json
import logging
import re
import os
import uuid
import csv
import secrets
import openpyxl
import asyncio
from datetime import datetime, timezone
from fastapi import APIRouter, Request, Form, HTTPException, UploadFile, File, BackgroundTasks
from fastapi.responses import HTMLResponse, RedirectResponse, JSONResponse

from app import db, config, auth, secure_store, meta_provider, prompt_assistant, skill_runtime, calendar_client, file_parser

log = logging.getLogger("client-panel")
router = APIRouter(prefix="/client", tags=["client-portal"])

# Reusable icons for the client dashboard sidebar and tabs
ICONS = {
    "dashboard": '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M3 13h8V3H3v10Z"/><path d="M13 21h8V11h-8v10Z"/><path d="M13 3h8v6h-8V3Z"/><path d="M3 21h8v-6H3v6Z"/></svg>',
    "wa": '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M22 11.08V12a10 10 0 1 1-5.93-9.14"/><polyline points="22 4 12 14.01 9 11.01"/></svg>',
    "prompt": '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M12 20h9"/><path d="M16.5 3.5a2.12 2.12 0 0 1 3 3L7 19l-4 1 1-4Z"/></svg>',
    "hours": '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><circle cx="12" cy="12" r="10"/><polyline points="12 6 12 12 16 14"/></svg>',
    "escalate": '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><polygon points="13 2 3 14 12 14 11 22 21 10 12 10 13 2"/></svg>',
    "knowledge": '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M14 2H6a2 2 0 0 0-2 2v16a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V8z"/><polyline points="14 2 14 8 20 8"/><line x1="16" y1="13" x2="8" y2="13"/><line x1="16" y1="17" x2="8" y2="17"/><polyline points="10 9 9 9 8 9"/></svg>',
    "integrations": '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><rect x="3" y="3" width="7" height="9"/><rect x="14" y="3" width="7" height="5"/><rect x="14" y="12" width="7" height="9"/><rect x="3" y="16" width="7" height="5"/></svg>',
    "out": '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M9 21H5a2 2 0 0 1-2-2V5a2 2 0 0 1 2-2h4"/><path d="M16 17l5-5-5-5"/><path d="M21 12H9"/></svg>',
    "success": '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.5" class="green"><path d="M22 11.08V12a10 10 0 1 1-5.93-9.14"/><polyline points="22 4 12 14.01 9 11.01"/></svg>',
    "error": '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.5" class="red"><circle cx="12" cy="12" r="10"/><line x1="15" y1="9" x2="9" y2="15"/><line x1="9" y1="9" x2="15" y2="15"/></svg>',
    "chat": '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M21 15a4 4 0 0 1-4 4H8l-5 3V7a4 4 0 0 1 4-4h10a4 4 0 0 1 4 4v8Z"/></svg>',
    "leads": '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M16 21v-2a4 4 0 0 0-4-4H6a4 4 0 0 0-4 4v2"/><circle cx="9" cy="7" r="4"/><path d="M22 21v-2a4 4 0 0 0-3-3.87"/><path d="M16 3.13a4 4 0 0 1 0 7.75"/></svg>',
    "settings": '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M12.22 2h-.44a2 2 0 0 0-2 2v.18a2 2 0 0 1-1 1.73l-.43.25a2 2 0 0 1-2 0l-.15-.08a2 2 0 0 0-2.73.73l-.22.38a2 2 0 0 0 .73 2.73l.15.1a2 2 0 0 1 1 1.72v.51a2 2 0 0 1-1 1.74l-.15.09a2 2 0 0 0-.73 2.73l.22.38a2 2 0 0 0 2.73.73l.15-.08a2 2 0 0 1 2 0l.43.25a2 2 0 0 1 1 1.73V20a2 2 0 0 0 2 2h.44a2 2 0 0 0 2-2v-.18a2 2 0 0 1 1-1.73l.43-.25a2 2 0 0 1 2 0l.15.08a2 2 0 0 0 2.73-.73l.22-.39a2 2 0 0 0-.73-2.73l-.15-.08a2 2 0 0 1-1-1.74v-.5a2 2 0 0 1 1-1.74l.15-.09a2 2 0 0 0 .73-2.73l-.22-.38a2 2 0 0 0-2.73-.73l-.15.08a2 2 0 0 1-2 0l-.43-.25a2 2 0 0 1-1-1.73V4a2 2 0 0 0-2-2z"/><circle cx="12" cy="12" r="3"/></svg>',
    "templates": '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><rect x="3" y="3" width="18" height="18" rx="2" ry="2"/><line x1="9" y1="9" x2="15" y2="9"/><line x1="9" y1="13" x2="15" y2="13"/><line x1="9" y1="17" x2="13" y2="17"/></svg>',
    "contacts": '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M4 19.5A2.5 2.5 0 0 1 6.5 17H20"/><path d="M6.5 2H20v20H6.5A2.5 2.5 0 0 1 4 19.5v-15A2.5 2.5 0 0 1 6.5 2z"/><circle cx="12" cy="8" r="2"/><path d="M12 11c-2 0-3 1-3 2v1h6v-1c0-1-1-2-3-2Z"/></svg>',
    "campaigns": '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M18 8A6 6 0 0 0 6 8c0 7-3 9-3 9h18s-3-2-3-9Z"/><path d="M13.73 21a2 2 0 0 1-3.46 0"/><path d="M2 10h4M18 10h4M4 5l2.5 2.5M17.5 7.5L20 5"/></svg>'
}

CLIENT_CSS = """
<style>
  @import url('https://fonts.googleapis.com/css2?family=Outfit:wght@300;400;500;600;700;800&family=Inter:wght@300;400;500;600;700&display=swap');

  :root {
    --bg: #f8fafc;
    --panel: #ffffff;
    --ink: #0f172a;
    --muted: #64748b;
    --line: #e2e8f0;
    --line-strong: #cbd5e1;
    --primary: #0d9488;
    --primary-dark: #0f766e;
    --primary-light: #ccfbf1;
    --amber: #d97706;
    --red: #e11d48;
    --blue: #2563eb;
    --green: #059669;
    --shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.05), 0 2px 4px -2px rgba(0, 0, 0, 0.05);
    --font: 'Inter', system-ui, sans-serif;
    --font-display: 'Outfit', sans-serif;
    --transition: all 0.2s cubic-bezier(0.4, 0, 0.2, 1);
  }

  * { box-sizing: border-box; }
  body {
    margin: 0;
    background: var(--bg);
    color: var(--ink);
    font-family: var(--font);
    -webkit-font-smoothing: antialiased;
  }
  
  .layout {
    display: grid;
    grid-template-columns: 260px 1fr;
    min-height: 100vh;
  }

  /* Sidebar styling */
  .sidebar {
    background: #0f172a;
    color: #e2e8f0;
    padding: 24px 18px;
    display: flex;
    flex-direction: column;
    justify-content: space-between;
    height: 100vh;
    position: sticky;
    top: 0;
    z-index: 50;
    border-right: 1px solid rgba(255, 255, 255, 0.05);
  }
  .brand {
    display: flex;
    align-items: center;
    gap: 12px;
    margin-bottom: 32px;
  }
  .brand-logo {
    width: 38px;
    height: 38px;
    background: var(--primary);
    color: white;
    font-family: var(--font-display);
    font-weight: 800;
    font-size: 16px;
    border-radius: 10px;
    display: grid;
    place-items: center;
    box-shadow: 0 4px 12px rgba(13, 148, 136, 0.3);
  }
  .brand-text strong {
    display: block;
    font-size: 16px;
    color: white;
    font-family: var(--font-display);
  }
  .brand-text span {
    font-size: 11px;
    color: #94a3b8;
  }
  .sidebar-nav {
    display: flex;
    flex-direction: column;
    gap: 6px;
    flex-grow: 1;
  }
  .sidebar-link {
    display: flex;
    align-items: center;
    gap: 12px;
    padding: 11px 14px;
    color: #94a3b8;
    font-size: 14px;
    font-weight: 500;
    border-radius: 8px;
    cursor: pointer;
    transition: var(--transition);
  }
  .sidebar-link:hover {
    background: rgba(255, 255, 255, 0.05);
    color: white;
  }
  .sidebar-link.active {
    background: var(--primary);
    color: white;
    font-weight: 600;
    box-shadow: 0 4px 12px rgba(13, 148, 136, 0.25);
  }
  .sidebar-link svg {
    width: 18px;
    height: 18px;
  }
  .sidebar-user-profile {
    margin-top: auto;
    display: flex;
    align-items: center;
    gap: 10px;
    padding: 12px;
    background: rgba(255, 255, 255, 0.03);
    border: 1px solid rgba(255, 255, 255, 0.08);
    border-radius: 10px;
  }
  .user-avatar {
    width: 32px;
    height: 32px;
    border-radius: 50%;
    background: var(--primary);
    color: white;
    display: flex;
    align-items: center;
    justify-content: center;
    font-weight: 700;
    font-size: 13px;
    flex-shrink: 0;
  }
  .user-info {
    flex: 1;
    display: flex;
    flex-direction: column;
    overflow: hidden;
  }
  .user-info strong {
    color: white;
    font-size: 13px;
    white-space: nowrap;
    overflow: hidden;
    text-overflow: ellipsis;
  }
  .user-info span {
    color: #94a3b8;
    font-size: 11px;
  }
  .icon-btn {
    background: none;
    border: none;
    color: #94a3b8;
    cursor: pointer;
    padding: 4px;
    display: flex;
    align-items: center;
    justify-content: center;
    transition: var(--transition);
    border-radius: 6px;
  }
  .icon-btn svg {
    width: 16px;
    height: 16px;
  }
  .icon-btn:hover {
    color: white;
    background: rgba(255, 255, 255, 0.1);
  }
  .icon-btn.logout:hover {
    color: var(--red);
    background: rgba(239, 68, 68, 0.1);
  }
  
  /* Main Content Area */
  .main-content {
    padding: 30px 40px;
    overflow-y: auto;
  }
  .header {
    display: flex;
    justify-content: space-between;
    align-items: flex-start;
    margin-bottom: 28px;
  }
  .header h1 {
    font-family: var(--font-display);
    font-size: 28px;
    font-weight: 800;
    margin: 0;
    color: var(--ink);
  }
  .header p {
    margin: 4px 0 0;
    color: var(--muted);
    font-size: 14px;
  }
  .header-actions {
    display: flex;
    align-items: center;
    gap: 12px;
  }
  
  /* Badges & Widgets */
  .badge {
    display: inline-flex;
    align-items: center;
    padding: 4px 10px;
    border-radius: 9999px;
    font-size: 12px;
    font-weight: 600;
    text-transform: uppercase;
    letter-spacing: 0.02em;
    background: #e2e8f0;
    color: #475569;
  }
  .badge.success { background: #d1fae5; color: #065f46; }
  .badge.warning { background: #fef3c7; color: #b45309; }
  .badge.danger { background: #ffe4e6; color: #9f1239; }
  .badge-toggle {
    transition: opacity 0.2s ease, transform 0.1s ease;
  }
  .badge-toggle:hover {
    opacity: 0.85;
    transform: translateY(-1px);
  }
  .badge-toggle:active {
    transform: translateY(0);
  }
  
  /* Tab Panel */
  .tab-panel {
    display: none;
    animation: fadeIn 0.2s ease-in-out;
  }
  .tab-panel.active {
    display: block;
  }
  @keyframes fadeIn {
    from { opacity: 0; transform: translateY(4px); }
    to { opacity: 1; transform: translateY(0); }
  }

  /* Grid Layouts */
  .grid-2 {
    display: grid;
    grid-template-columns: repeat(2, 1fr);
    gap: 24px;
  }
  .grid-3 {
    display: grid;
    grid-template-columns: repeat(3, 1fr);
    gap: 20px;
  }
  .grid-cards {
    display: grid;
    grid-template-columns: minmax(0, 1.4fr) minmax(280px, 0.8fr);
    gap: 24px;
  }

  /* UI Cards */
  .card {
    background: var(--panel);
    border: 1px solid var(--line);
    border-radius: 12px;
    box-shadow: var(--shadow);
    padding: 24px;
    margin-bottom: 24px;
  }
  .card-header {
    margin-bottom: 18px;
  }
  .card-header h2 {
    font-family: var(--font-display);
    font-size: 18px;
    font-weight: 700;
    margin: 0;
    color: var(--ink);
  }
  .card-header p {
    font-size: 13px;
    color: var(--muted);
    margin: 4px 0 0;
  }
  
  /* KPI Widget */
  .kpi-card {
    background: white;
    border: 1px solid var(--line);
    border-radius: 12px;
    padding: 20px;
    box-shadow: var(--shadow);
    position: relative;
    overflow: hidden;
  }
  .kpi-title {
    font-size: 12px;
    font-weight: 600;
    color: var(--muted);
    text-transform: uppercase;
    letter-spacing: 0.05em;
  }
  .kpi-value {
    font-family: var(--font-display);
    font-size: 28px;
    font-weight: 800;
    margin-top: 10px;
    color: var(--ink);
  }
  .kpi-status {
    font-size: 12px;
    margin-top: 6px;
    display: flex;
    align-items: center;
    gap: 4px;
  }
  
  /* Form elements */
  label {
    display: block;
    font-size: 13px;
    font-weight: 600;
    color: var(--ink);
    margin: 14px 0 6px;
  }
  input:not([type="checkbox"]):not([type="radio"]):not([type="submit"]):not([type="button"]):not([type="hidden"]), select, textarea {
    width: 100%;
    border: 1px solid var(--line-strong);
    border-radius: 8px;
    padding: 11px 13px;
    background: white;
    color: var(--ink);
    transition: var(--transition);
    outline: none;
  }
  input:not([type="checkbox"]):not([type="radio"]):not([type="submit"]):not([type="button"]):not([type="hidden"]):focus, select:focus, textarea:focus {
    border-color: var(--primary);
    box-shadow: 0 0 0 3px rgba(13, 148, 136, 0.15);
  }
  .checkbox-group {
    display: flex;
    align-items: center;
    gap: 10px;
    margin: 12px 0;
  }
  .checkbox-group input {
    width: 18px;
    height: 18px;
    accent-color: var(--primary);
  }
  
  /* Buttons */
  .btn {
    display: inline-flex;
    align-items: center;
    justify-content: center;
    gap: 8px;
    background: var(--ink);
    color: white;
    border: 1px solid var(--ink);
    border-radius: 8px;
    padding: 10px 18px;
    font-weight: 600;
    font-size: 13px;
    cursor: pointer;
    transition: var(--transition);
  }
  .btn:hover {
    background: #1e293b;
    border-color: #1e293b;
    transform: translateY(-1px);
    box-shadow: 0 4px 6px rgba(0,0,0,0.06);
  }
  .btn:active {
    transform: translateY(0);
  }
  .btn.primary-btn {
    background: var(--primary);
    border-color: var(--primary);
    box-shadow: 0 4px 10px rgba(13, 148, 136, 0.2);
  }
  .btn.primary-btn:hover {
    background: var(--primary-dark);
    border-color: var(--primary-dark);
  }
  .btn.secondary {
    background: white;
    color: var(--ink);
    border-color: var(--line-strong);
  }
  .btn.secondary:hover {
    background: #f8fafc;
  }
  .btn.whatsapp-btn {
    background: #128c7e;
    border-color: #128c7e;
    box-shadow: 0 4px 10px rgba(18, 140, 126, 0.2);
  }
  .btn.whatsapp-btn:hover {
    background: #075e54;
    border-color: #075e54;
  }
  .btn:disabled {
    opacity: 0.5;
    cursor: not-allowed;
    transform: none !important;
    box-shadow: none !important;
  }

  /* Table styling */
  table {
    width: 100%;
    border-collapse: collapse;
  }
  th, td {
    padding: 12px 16px;
    border-bottom: 1px solid var(--line);
    text-align: left;
  }
  th {
    background: #f8fafc;
    color: var(--muted);
    font-size: 11px;
    font-weight: 700;
    text-transform: uppercase;
    letter-spacing: 0.05em;
  }
  tr:hover td {
    background: rgba(13, 148, 136, 0.01);
  }
  
  /* Alert banners */
  .notice-banner {
    padding: 12px 16px;
    border-radius: 8px;
    font-size: 13px;
    font-weight: 500;
    margin-bottom: 20px;
    display: flex;
    align-items: center;
    gap: 8px;
    border: 1px solid transparent;
  }
  .notice-banner svg {
    width: 20px;
    height: 20px;
    flex-shrink: 0;
  }
  .notice-banner.success {
    background: #d1fae5;
    border-color: #a7f3d0;
    color: #065f46;
  }
  .notice-banner.error {
    background: #ffe4e6;
    border-color: #fecdd3;
    color: #9f1239;
  }
  
  /* Helper Classes */
  .muted-text { color: var(--muted); font-size: 13px; }
  .bold-text { font-weight: 600; }
  .sync-status { font-size: 12px; font-weight: 500; color: var(--muted); margin-top: 8px; }
  .sync-status.ok { color: var(--green); }
  .sync-status.err { color: var(--red); }
  
  /* Weekday Grid for Schedules */
  .weekday-row {
    display: grid;
    grid-template-columns: 140px 110px 1fr 1fr;
    gap: 16px;
    align-items: center;
    padding: 12px 8px;
    border-bottom: 1px solid var(--line);
  }
  .weekday-row:last-child {
    border-bottom: 0;
  }
  
  /* Prompt Assistant block */
  .prompt-workspace {
    display: grid;
    grid-template-columns: 1fr 1.2fr;
    gap: 20px;
  }
  .chat-bubble-container {
    max-height: 380px;
    overflow-y: auto;
    border: 1px solid var(--line);
    border-radius: 8px;
    padding: 12px;
    background: #f8fafc;
  }
  .bubble {
    max-width: 85%;
    padding: 10px 12px;
    border-radius: 8px;
    margin-bottom: 8px;
    font-size: 13px;
    line-height: 1.4;
  }
  .bubble.user {
    background: white;
    border: 1px solid var(--line);
  }
  .bubble.assistant {
    background: var(--primary-light);
    color: var(--primary-dark);
    margin-left: auto;
    border: 1px solid rgba(13, 148, 136, 0.1);
  }
  .bubble-meta {
    font-size: 10px;
    color: var(--muted);
    margin-bottom: 4px;
  }

  @media (max-width: 1024px) {
    .layout { grid-template-columns: 1fr; }
    .sidebar { height: auto; position: static; }
    .grid-2, .grid-cards, .prompt-workspace { grid-template-columns: 1fr; }
  }
  .password-wrapper {
    position: relative;
    display: flex;
    align-items: center;
    width: 100%;
  }
  .password-wrapper input {
    padding-right: 40px !important;
  }
  .password-toggle {
    position: absolute;
    right: 10px;
    background: none;
    border: none;
    cursor: pointer;
    color: var(--muted);
    padding: 0;
    display: flex;
    align-items: center;
    justify-content: center;
    transition: var(--transition);
    z-index: 10;
  }
  .password-toggle:hover {
    color: var(--primary);
  }

  /* Scrollbar styling */
  ::-webkit-scrollbar { width: 6px; height: 6px; }
  ::-webkit-scrollbar-track { background: transparent; }
  ::-webkit-scrollbar-thumb { background: #cbd5e1; border-radius: 999px; }
  ::-webkit-scrollbar-thumb:hover { background: #94a3b8; }

  /* Chatwoot Layout */
  .chatwoot-layout {
    display: grid;
    grid-template-columns: 300px 1fr 280px;
    height: calc(100vh - 220px);
    background: white;
    border: 1px solid var(--line);
    border-radius: 12px;
    overflow: hidden;
    box-shadow: var(--shadow);
    margin-top: 14px;
  }
  .chat-sidebar, .chat-main, .chat-crm {
    min-height: 0;
  }
  .chat-sidebar {
    border-right: 1px solid var(--line);
    background: #f8fafc;
    display: flex;
    flex-direction: column;
  }
  .chat-sidebar-header {
    padding: 16px;
    border-bottom: 1px solid var(--line);
    font-weight: 700;
    font-size: 15px;
    display: flex;
    justify-content: space-between;
    align-items: center;
  }
  .chat-list {
    overflow-y: auto;
    flex: 1;
  }
  .chat-item {
    padding: 14px 16px;
    border-bottom: 1px solid var(--line);
    cursor: pointer;
    transition: var(--transition);
    text-decoration: none;
    display: block;
    color: inherit;
  }
  .chat-item:hover { background: rgba(13, 148, 136, 0.05); }
  .chat-item.active { background: white; border-left: 3px solid var(--primary); }
  .chat-item-header { display: flex; justify-content: space-between; margin-bottom: 4px; align-items: center; }
  .chat-item-name { font-weight: 600; color: var(--ink); font-size: 14px; }
  .chat-item-time { font-size: 11px; color: var(--muted); }
  .chat-item-preview { font-size: 12px; color: var(--muted); white-space: nowrap; overflow: hidden; text-overflow: ellipsis; }
  
  .chat-main {
    display: flex;
    flex-direction: column;
    background: white;
  }
  .chat-header {
    padding: 16px 20px;
    border-bottom: 1px solid var(--line);
    display: flex;
    align-items: center;
    gap: 12px;
  }
  .chat-avatar {
    width: 36px; height: 36px; border-radius: 50%; background: var(--primary-light); color: var(--primary-dark);
    display: flex; align-items: center; justify-content: center; font-weight: 700; font-size: 14px;
  }
  .chat-header-info { display: flex; flex-direction: column; }
  .chat-header-info strong { font-size: 15px; }
  .chat-header-info small { color: var(--muted); font-size: 12px; }
  .chat-messages {
    flex: 1;
    overflow-y: auto;
    padding: 20px;
    display: flex;
    flex-direction: column;
    gap: 12px;
    background: #f8fafc;
  }
  .chat-crm {
    border-left: 1px solid var(--line);
    background: white;
    padding: 20px;
    overflow-y: auto;
  }
  .crm-section { margin-bottom: 24px; }
  .crm-section h3 { font-size: 11px; text-transform: uppercase; color: var(--muted); font-weight: 700; margin: 0 0 12px 0; letter-spacing: 0.05em; }

  .bubble {
    max-width: min(720px, 85%);
    padding: 12px 14px;
    border-radius: 12px;
    border: 1px solid var(--line);
    background: white;
    box-shadow: 0 2px 4px rgba(0,0,0,0.02);
  }
  .bubble.assistant { margin-left: auto; background: #ccfbf1; border-color: #99f6e4; color: var(--primary-dark); }
  .bubble .meta { font-size: 11px; color: var(--muted); margin-bottom: 6px; font-weight: 600; text-transform: uppercase; letter-spacing: 0.02em; }

  .tabs { display: flex; gap: 8px; flex-wrap: wrap; margin-bottom: 14px; }
  .tabs a {
    border: 1px solid var(--line);
    background: white;
    border-radius: 8px;
    padding: 8px 12px;
    font-size: 13px;
    font-weight: 500;
    color: var(--muted);
    text-decoration: none;
    transition: var(--transition);
  }
  .tabs a:hover { border-color: var(--line-strong); color: var(--ink); }
  .tabs a.active { background: var(--ink); border-color: var(--ink); color: white; font-weight: 600; }

  .b-en_progreso, .b-pendiente { background: #fef3c7; color: #b45309; }
  .b-calificado, .b-resuelto { background: #d1fae5; color: #065f46; }
  .b-descalificado, .b-urgente { background: #ffe4e6; color: #9f1239; }

  /* Modal overlay and content styles */
  .modal-overlay {
    position: fixed;
    top: 0;
    left: 0;
    width: 100%;
    height: 100%;
    background: rgba(15, 23, 42, 0.6);
    backdrop-filter: blur(4px);
    display: none;
    justify-content: center;
    align-items: center;
    z-index: 1000;
    opacity: 0;
    transition: opacity 0.2s ease-in-out;
  }
  .modal-overlay.active {
    display: flex;
    opacity: 1;
  }
  .modal-content {
    background: white;
    border-radius: 12px;
    width: 90%;
    max-width: 550px;
    box-shadow: 0 20px 25px -5px rgba(0, 0, 0, 0.1), 0 10px 10px -5px rgba(0, 0, 0, 0.04);
    overflow: hidden;
    transform: translateY(20px);
    transition: transform 0.2s ease-in-out;
  }
  .modal-overlay.active .modal-content {
    transform: translateY(0);
  }
  .modal-header {
    padding: 16px 20px;
    border-bottom: 1px solid var(--line);
    display: flex;
    justify-content: space-between;
    align-items: center;
  }
  .modal-header h3 {
    margin: 0;
    font-family: var(--font-display);
    font-weight: 700;
    font-size: 18px;
    color: var(--ink);
  }
  .modal-body {
    padding: 20px;
  }
  .modal-footer {
    padding: 12px 20px;
    background: #f8fafc;
    border-top: 1px solid var(--line);
    display: flex;
    justify-content: flex-end;
    gap: 8px;
  }
  .modal-close-btn {
    background: none;
    border: none;
    font-size: 24px;
    cursor: pointer;
    color: var(--muted);
    line-height: 1;
    padding: 0;
  }
  .modal-close-btn:hover {
    color: var(--ink);
  }
</style>
"""

def _require_client_login(request: Request) -> dict:
    user = request.session.get("user")
    if not user:
        raise HTTPException(status_code=302, headers={"Location": "/admin/login"})
    csrf_token = request.session.setdefault("_csrf_token", secrets.token_urlsafe(32))
    session = {
        "user": user,
        "role": request.session.get("role", "client_viewer"),
        "client_id": request.session.get("client_id"),
        "user_id": request.session.get("user_id"),
        "name": request.session.get("name") or user,
        "_csrf_token": csrf_token,
    }
    if session["role"] == "agency_admin" and not session["client_id"]:
        # Superadmin preview: fallback to first client if none in session
        pass
    elif not session["client_id"]:
        raise HTTPException(status_code=403, detail="Sesión inválida: sin cliente asignado.")
    return session

async def _require_bot_access(session: dict, bot_id: int) -> dict:
    bot = await db.get_bot(bot_id)
    if not bot:
        raise HTTPException(status_code=404, detail="Bot no encontrado")
    if session.get("role") != "agency_admin" and bot.get("client_id") != session.get("client_id"):
        raise HTTPException(status_code=403, detail="Sin acceso a este bot")
    return bot

async def _require_bot_editor(session: dict, bot_id: int) -> dict:
    bot = await _require_bot_access(session, bot_id)
    if session.get("role") in ("agency_admin", "client_admin"):
        return bot
    raise HTTPException(status_code=403, detail="Tu rol es de solo lectura. No puedes modificar este bot.")

async def _first_allowed_bot(session: dict) -> dict | None:
    client_id = session.get("client_id")
    bots = await db.list_bots(client_id=client_id, limit=1)
    return bots[0] if bots else None

def _fmt_dt(value: datetime | None) -> str:
    return value.strftime("%d/%m/%Y %H:%M") if value else "-"


def _clip(value: str | None, size: int = 160) -> str:
    text = (value or "").strip()
    return text if len(text) <= size else text[: size - 1].rstrip() + "..."


def _wa_link(wa_id: str) -> str:
    return f"https://wa.me/{html.escape(wa_id)}"


def _display_name(name: str | None, wa_id: str) -> str:
    clean = (name or "").strip()
    invalid_terms = (
        "quien", "quien toma", "toma", "decision", "decisión", "dueno",
        "dueño", "negocio", "empresa", "encargado", "responsable",
    )
    if not clean or any(term in clean.lower() for term in invalid_terms):
        return wa_id
    return clean


def _layout(title: str, body: str, session: dict, active_tab: str = "inicio", notice: str = "", bots_list: list = [], selected_bot_id: int | None = None, chatwoot_enabled: bool = False, cw_base_url: str = "", cw_account: str = "") -> str:
    csrf_token = session.setdefault("_csrf_token", secrets.token_urlsafe(32))
    bot_options = "".join(
        f'<option value="{b["id"]}" {"selected" if b["id"] == selected_bot_id else ""}>{html.escape(b["name"])}</option>'
        for b in bots_list
    )
    
    sidebar_links = [
        ("inicio", "Inicio", ICONS["dashboard"]),
        ("conversations", "Conversaciones", ICONS["chat"]),
        ("crm", "CRM de Leads", ICONS["leads"]),
        ("contacts", "Contactos", ICONS["contacts"]),
        ("campaigns", "Campañas", ICONS["campaigns"]),
        ("whatsapp", "Conectar WhatsApp", ICONS["wa"]),
        ("prompt", "Comportamiento (IA)", ICONS["prompt"]),
        ("hours", "Horarios", ICONS["hours"]),
        ("escalate", "Reglas de Escalado", ICONS["escalate"]),
        ("knowledge", "Base de Conocimiento", ICONS["knowledge"]),
        ("templates", "Plantillas", ICONS["templates"]),
        ("integrations", "Integraciones", ICONS["integrations"]),
    ]
    
    links_html = ""
    for key, label, icon in sidebar_links:
        if key == "conversations" and chatwoot_enabled and cw_base_url and cw_account:
            cw_url = f"{cw_base_url.rstrip('/')}/app/accounts/{cw_account}/dashboard"
            links_html += f'<a href="{cw_url}" target="_blank" class="sidebar-link" style="text-decoration: none;">{icon}<span>Chatwoot ↗</span></a>'
        else:
            links_html += f'<div class="sidebar-link {"active" if key == active_tab else ""}" onclick="switchTab(\'{key}\')">{icon}<span>{label}</span></div>'
    
    selector_html = ""
    if len(bots_list) > 1:
        selector_html = f"""
        <div style="margin-bottom: 20px;">
          <label style="color:#94a3b8; font-size:12px; margin-top:0;">Seleccionar Bot</label>
          <select id="sidebarBotSelector" onchange="changeActiveBot(this.value)" style="background:#1e293b; color:white; border-color:rgba(255,255,255,0.1); font-size:13px; padding:8px 10px;">
            {bot_options}
          </select>
        </div>
        """

    return f"""<!doctype html>
<html lang="es">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{html.escape(title)} - Asistto</title>
  {CLIENT_CSS}
  <script>
    function togglePasswordVisibility(btn) {{
      const input = btn.previousElementSibling;
      if (input.type === "password") {{
        input.type = "text";
        btn.innerHTML = `<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round" style="width: 18px; height: 18px;"><path d="M17.94 17.94A10.07 10.07 0 0 1 12 20c-7 0-11-8-11-8a18.45 18.45 0 0 1 5.06-5.94M9.9 4.24A9.12 9.12 0 0 1 12 4c7 0 11 8 11 8a18.5 18.5 0 0 1-2.16 3.19m-6.72-1.07a3 3 0 1 1-4.24-4.24"></path><line x1="1" y1="1" x2="23" y2="23"></line></svg>`;
      }} else {{
        input.type = "password";
        btn.innerHTML = `<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round" style="width: 18px; height: 18px;"><path d="M1 12s4-8 11-8 11 8 11 8-4 8-11 8-11-8-11-8z"></path><circle cx="12" cy="12" r="3"></circle></svg>`;
      }}
    }}
    document.addEventListener("submit", function(event) {{
      const form = event.target;
      if (!form || String(form.method || "").toLowerCase() !== "post") return;
      try {{
        const actionUrl = new URL(form.action || window.location.href);
        actionUrl.searchParams.set("csrf_token", "{html.escape(csrf_token)}");
        form.action = actionUrl.pathname + actionUrl.search;
      }} catch (e) {{}}
      if (form.querySelector('input[name="csrf_token"]')) return;
      const input = document.createElement("input");
      input.type = "hidden";
      input.name = "csrf_token";
      input.value = "{html.escape(csrf_token)}";
      form.appendChild(input);
    }});
  </script>
</head>
<body>
  <div class="layout">
    <aside class="sidebar">
      <div>
        <div class="brand">
          <div class="brand-logo">AH</div>
          <div class="brand-text">
            <strong>Asistto</strong>
            <span>by Humanio</span>
          </div>
        </div>
        <div style="padding: 0 16px; margin: -10px 0 16px; font-size: 11px; color: #94a3b8;">
          ID del Bot: <strong style="color: #38bdf8;">{selected_bot_id}</strong>
        </div>
        {selector_html}
        <nav class="sidebar-nav">
          {links_html}
        </nav>
      </div>
      <div class="sidebar-user-profile">
        <div class="user-avatar">{html.escape(session.get("name", "Usuario")[:2].upper())}</div>
        <div class="user-info" title="{html.escape(session.get("name", "Usuario"))}">
          <strong>{html.escape(session.get("name", "Usuario"))}</strong>
          <span style="display:flex; align-items:center; gap:4px; margin-top:2px;">
             Configuración <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" style="width:12px; height:12px;"><path d="M12.22 2h-.44a2 2 0 0 0-2 2v.18a2 2 0 0 1-1 1.73l-.43.25a2 2 0 0 1-2 0l-.15-.08a2 2 0 0 0-2.73.73l-.22.38a2 2 0 0 0 .73 2.73l.15.1a2 2 0 0 1 1 1.72v.51a2 2 0 0 1-1 1.74l-.15.09a2 2 0 0 0-.73 2.73l.22.38a2 2 0 0 0 2.73.73l.15-.08a2 2 0 0 1 2 0l.43.25a2 2 0 0 1 1 1.73V20a2 2 0 0 0 2 2h.44a2 2 0 0 0 2-2v-.18a2 2 0 0 1 1-1.73l.43-.25a2 2 0 0 1 2 0l.15.08a2 2 0 0 0 2.73-.73l.22-.39a2 2 0 0 0-.73-2.73l-.15-.08a2 2 0 0 1-1-1.74v-.5a2 2 0 0 1 1-1.74l.15-.09a2 2 0 0 0 .73-2.73l-.22-.38a2 2 0 0 0-2.73-.73l-.15.08a2 2 0 0 1-2 0l-.43-.25a2 2 0 0 1-1-1.73V4a2 2 0 0 0-2-2z"></path><circle cx="12" cy="12" r="3"></circle></svg>
          </span>
        </div>
        <form method="post" action="/admin/logout" style="margin: 0; display: flex;">
          <input type="hidden" name="csrf_token" value="{html.escape(csrf_token)}">
          <button class="icon-btn logout" type="submit" title="Cerrar sesión" style="color:var(--red);"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" style="width:18px;height:18px;"><path d="M9 21H5a2 2 0 0 1-2-2V5a2 2 0 0 1 2-2h4"></path><polyline points="16 17 21 12 16 7"></polyline><line x1="21" y1="12" x2="9" y2="12"></line></svg></button>
        </form>
      </div>
    </aside>
    <main class="main-content">
      {notice}
      {body}
    </main>
  </div>
  
  <script>
    function switchTab(tabId) {{
      // Update sidebar links
      document.querySelectorAll('.sidebar-link').forEach(link => {{
        link.classList.remove('active');
        const onclickAttr = link.getAttribute('onclick');
        if (onclickAttr && onclickAttr.includes(tabId)) {{
          link.classList.add('active');
        }}
      }});
      // Update tab panels
      document.querySelectorAll('.tab-panel').forEach(panel => {{
        panel.classList.remove('active');
      }});
      const activePanel = document.getElementById('panel-' + tabId);
      if (activePanel) {{
        activePanel.classList.add('active');
      }}
      
      // Auto-scroll chat messages if switching to conversations tab
      if (tabId === 'conversations') {{
        const chatMessages = document.querySelector('.chat-messages');
        if (chatMessages) {{
          chatMessages.scrollTop = chatMessages.scrollHeight;
        }}
      }}
      
      // Save tab to URL state
      const url = new URL(window.location.href);
      url.searchParams.set("tab", tabId);
      window.history.replaceState(null, "", url);
    }}
    
    function changeActiveBot(botId) {{
      const url = new URL(window.location.href);
      url.searchParams.set("bot_id", botId);
      window.location.href = url.toString();
    }}
    
    // Auto-activate tab from URL parameters
    window.addEventListener('DOMContentLoaded', () => {{
      const params = new URLSearchParams(window.location.search);
      const tab = params.get("tab") || "inicio";
      switchTab(tab);
      
      // Auto-scroll chat messages on load
      const chatMessages = document.querySelector('.chat-messages');
      if (chatMessages) {{
        chatMessages.scrollTop = chatMessages.scrollHeight;
      }}
    }});

    // Auto-refresh chat list and messages in background (Conversations tab)
    setInterval(async () => {{
      if (document.hidden) return;
      
      const chatLayout = document.querySelector('.chatwoot-layout');
      if (!chatLayout || chatLayout.offsetParent === null) return;
      
      try {{
        const response = await fetch(window.location.href);
        if (response.ok) {{
          const html = await response.text();
          const parser = new DOMParser();
          const doc = parser.parseFromString(html, 'text/html');
          
          const newChatList = doc.querySelector('.chat-list');
          const currentChatList = document.querySelector('.chat-list');
          if (newChatList && currentChatList) {{
            if (currentChatList.innerHTML !== newChatList.innerHTML) {{
              currentChatList.innerHTML = newChatList.innerHTML;
            }}
          }}
          
          const newMessages = doc.querySelector('.chat-messages');
          const currentMessages = document.querySelector('.chat-messages');
          if (newMessages && currentMessages) {{
            if (currentMessages.innerHTML !== newMessages.innerHTML) {{
              const hadNewMessages = newMessages.children.length !== currentMessages.children.length;
              const wasAtBottom = currentMessages.scrollHeight - currentMessages.scrollTop <= currentMessages.clientHeight + 100;
              
              currentMessages.innerHTML = newMessages.innerHTML;
              
              if (hadNewMessages && wasAtBottom) {{
                currentMessages.scrollTop = currentMessages.scrollHeight;
              }}
            }}
          }}
          
          const newCrm = doc.querySelector('.chat-crm');
          const currentCrm = document.querySelector('.chat-crm');
          if (newCrm && currentCrm) {{
            if (currentCrm.innerHTML !== newCrm.innerHTML) {{
              currentCrm.innerHTML = newCrm.innerHTML;
            }}
          }}

          const newHeader = doc.querySelector('.chat-header');
          const currentHeader = document.querySelector('.chat-header');
          if (newHeader && currentHeader) {{
            if (currentHeader.innerHTML !== newHeader.innerHTML) {{
              currentHeader.innerHTML = newHeader.innerHTML;
            }}
          }}
        }}
      }} catch (err) {{
        console.error("Error refreshing conversations:", err);
      }}
    }}, 4000);
  </script>
</body>
</html>"""

@router.get("/app", response_class=HTMLResponse)
async def client_app(
    request: Request,
    bot_id: int | None = None,
    tab: str = "inicio",
    saved: str | None = None,
    wa_id: str | None = None,
    status: str = "en_progreso",
    search_q: str | None = None,
    tag_q: str | None = None,
    cpage: int = 1,
    map_file: str | None = None,
    imported: int | None = None,
    deleted: int | None = None,
    broadcast_selected: str | None = None,
):
    session = _require_client_login(request)
    client_id = session.get("client_id")
    
    # Query client's bots
    bots = await db.list_bots(client_id=client_id, limit=50)
    if not bots:
        csrf_token = session.get("_csrf_token") or ""
        return HTMLResponse(
            f"""<!doctype html>
            <html><head><title>Bienvenido a Asistto</title>{CLIENT_CSS}</head>
            <body style="background:#f8fafc; display:grid; place-items:center; min-height:100vh; padding:20px;">
              <div class="card" style="max-width:480px; text-align:center;">
                <div class="brand-logo" style="margin:0 auto 16px; width:48px; height:48px; font-size:18px;">AH</div>
                <h1 style="font-family:Outfit; font-size:24px; font-weight:800;">¡Bienvenido a Asistto!</h1>
                <p class="muted-text" style="margin-top:10px;">Tu cuenta está activa, pero aún no tienes ningún bot de WhatsApp configurado. Por favor, contacta a tu ejecutivo de cuenta para dar de alta tu primer bot y comenzar.</p>
                <form action="/admin/logout" method="post" style="margin-top:20px;">
                  <input type="hidden" name="csrf_token" value="{html.escape(csrf_token)}">
                  <button class="btn secondary" type="submit">Cerrar Sesión</button>
                </form>
              </div>
            </body></html>"""
        )
    
    selected_bot = None
    if bot_id:
        # Validate bot access
        try:
            selected_bot = await _require_bot_access(session, bot_id)
        except Exception:
            pass
            
    if not selected_bot:
        selected_bot = bots[0]
        bot_id = int(selected_bot["id"])
        
    # Notice banners
    notice_html = ""
    if saved == "1":
        notice_html = f'<div class="notice-banner success">{ICONS["success"]} Configuración guardada correctamente.</div>'
    elif saved == "err":
        notice_html = f'<div class="notice-banner error">{ICONS["error"]} Ocurrió un error. Revisa los datos.</div>'
    elif saved and saved.startswith("err_"):
        import urllib.parse
        decoded_msg = urllib.parse.unquote(saved[4:])
        notice_html = f'<div class="notice-banner error">{ICONS["error"]} {html.escape(decoded_msg)}</div>'
    elif saved == "err_parse":
        notice_html = f'<div class="notice-banner error">{ICONS["error"]} Error al leer o procesar el archivo. Asegúrate de que no esté dañado y sea de un tipo compatible (PDF, DOCX, MD, XLSX, CSV).</div>'
    elif saved == "err_ext":
        notice_html = f'<div class="notice-banner error">{ICONS["error"]} Formato de archivo invalido. Solo se admiten archivos .csv y .xlsx.</div>'
    elif saved == "err_num":
        notice_html = f'<div class="notice-banner error">{ICONS["error"]} El número de teléfono ingresado no es válido. Debe contener solo números y el prefijo de país.</div>'
    elif saved == "err_file_expired":
        notice_html = f'<div class="notice-banner error">{ICONS["error"]} El archivo temporal ha expirado o no se encuentra. Intenta subirlo de nuevo.</div>'
    elif saved == "err_file_too_large":
        notice_html = f'<div class="notice-banner error">{ICONS["error"]} El archivo excede el limite permitido para este panel.</div>'
    elif saved == "err_no_recipients":
        notice_html = f'<div class="notice-banner error">{ICONS["error"]} No hay destinatarios seleccionados o válidos para iniciar el envío masivo.</div>'
    elif saved == "err_confirm":
        notice_html = f'<div class="notice-banner error">{ICONS["error"]} Confirma el envio escribiendo CONFIRMAR antes de iniciar la campana.</div>'
    elif saved == "err_campaign_limit":
        notice_html = f'<div class="notice-banner error">{ICONS["error"]} La campana supera el limite de destinatarios permitido para este entorno.</div>'
    elif imported is not None:
        notice_html = f'<div class="notice-banner success">{ICONS["success"]} Se importaron con éxito {imported} contactos al directorio.</div>'
    elif deleted is not None:
        notice_html = f'<div class="notice-banner success">{ICONS["success"]} Se eliminaron {deleted} contactos correctamente.</div>'

    # 1. FETCH STATE
    # WhatsApp config
    wa_row = await db.get_bot_whatsapp_number(bot_id)
    wa_info = wa_row or {}
    
    # Check if access token is actually saved
    integration = await db.get_active_bot_integration(bot_id, "whatsapp_cloud")
    access_token_val = ""
    if integration:
        encrypted = await db.get_integration_secret_values(int(integration["id"]))
        for name in ("access_token", "whatsapp_access_token", "token"):
            enc_val = encrypted.get(name)
            if enc_val:
                dec = secure_store.decrypt_secret(enc_val)
                if dec:
                    access_token_val = dec
                    break
    has_token = bool(access_token_val)
    
    # Active prompt and PBD docs
    prompt_row = await db.get_active_bot_prompt(bot_id)
    current_prompt_content = (prompt_row.get("content") or "") if prompt_row else ""
    current_constitution = (prompt_row.get("pbd_constitution") or "") if prompt_row else ""
    current_specs = (prompt_row.get("pbd_specs") or "") if prompt_row else ""
    current_test_suite = (prompt_row.get("pbd_test_suite") or "") if prompt_row else ""
    
    # Business Hours Skill
    hours_skill = await db.get_bot_skill(bot_id, "business_hours")
    hours_config = (hours_skill.get("config") or {}) if hours_skill else {}
    hours_enabled = hours_skill.get("enabled", True) if hours_skill else True
    
    # Escalation Skill
    escalate_skill = await db.get_bot_skill(bot_id, "escalation")
    escalate_config = (escalate_skill.get("config") or {}) if escalate_skill else {}
    escalate_enabled = escalate_skill.get("enabled", True) if escalate_skill else True
    
    # Knowledge docs
    knowledge_docs = await db.list_bot_knowledge(bot_id, active_only=True)
    
    # Integrations
    calendar_status = await calendar_client.runtime_status(bot_id)
    calendar_config = {}
    calendar_integration = await db.get_active_bot_integration(bot_id, "google_calendar")
    if calendar_integration:
        calendar_config = calendar_integration.get("config") or {}
        
    api_integration = await db.get_active_bot_integration(bot_id, "external_api")
    api_config = {}
    api_secrets = []
    if api_integration:
        api_config = api_integration.get("config") or {}
        api_secrets = await db.list_integration_secrets(int(api_integration["id"]))
        
    chatwoot_integration = await db.get_active_bot_integration(bot_id, "chatwoot")
    chatwoot_config = {}
    chatwoot_secrets = {}
    chatwoot_enabled = False
    cw_account = ""
    cw_base_url = ""
    if chatwoot_integration:
        chatwoot_config = chatwoot_integration.get("config") or {}
        chatwoot_enabled = chatwoot_integration.get("enabled", False)
        # Fetch token secret
        enc_secrets = await db.get_integration_secret_values(int(chatwoot_integration["id"]))
        if "api_token" in enc_secrets:
            chatwoot_secrets["api_token"] = secure_store.decrypt_secret(enc_secrets["api_token"])
        
    # Routing Rules integration
    routing_integration = await db.get_bot_integration_by_type(bot_id, "routing_rules")
    routing_config = {}
    routing_enabled = False
    routing_secrets = {}
    if routing_integration:
        routing_config = routing_integration.get("config") or {}
        routing_enabled = routing_integration.get("enabled", False)
        enc_secrets = await db.get_integration_secret_values(int(routing_integration["id"]))
        if "webhook_auth_token" in enc_secrets:
            routing_secrets["webhook_auth_token"] = secure_store.decrypt_secret(enc_secrets["webhook_auth_token"])
            
    routing_webhook_url = routing_config.get("webhook_url", "")
    routing_phone_numbers = ",".join(routing_config.get("phone_numbers", []))
    routing_save_history = routing_config.get("save_history", False)
    routing_token_saved = bool(routing_secrets.get("webhook_auth_token"))
        
    env_rows_html = ""
    for sec in api_secrets:
        key_safe = html.escape(sec["secret_name"])
        env_rows_html += f'''
        <div class="env-var-row" style="display:flex; gap:8px; align-items:center; width:100%;">
          <input name="env_key" placeholder="KEY (Ej. STRIPE_KEY)" value="{key_safe}" style="flex:1; margin:0;" readonly>
          <div class="password-wrapper" style="flex:2;">
            <input type="password" name="env_val" placeholder="Valor del secreto" value="********" autocomplete="new-password" style="margin:0;">
            <button type="button" class="password-toggle" onclick="togglePasswordVisibility(this)">
              <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round" style="width: 18px; height: 18px;"><path d="M1 12s4-8 11-8 11 8 11 8-4 8-11 8-11-8-11-8z"></path><circle cx="12" cy="12" r="3"></circle></svg>
            </button>
          </div>
          <button type="button" class="btn secondary" style="padding:8px 12px; color:var(--red);" onclick="this.parentElement.remove()">X</button>
        </div>
        '''
        
    # Metrics
    metrics = await db.admin_metrics(bot_id=bot_id)
    
    # Recent conversation threads
    recent_threads = await db.list_conversation_threads(limit=5, bot_id=bot_id)
    
    # Fetch Meta templates
    templates = []
    templates_error = ""
    try:
        payload = await meta_provider.list_message_templates(bot_id)
        templates = payload.get("data") or []
    except Exception as exc:
        log.error(f"Error al listar plantillas para bot {bot_id}: {exc}")
        templates_error = str(exc)

    # Fetch Contacts and Campaigns
    search_contacts = search_q.strip() if search_q else None
    tag_filter = tag_q.strip() if tag_q else None
    contacts_limit = 50
    contacts_offset = max(0, cpage - 1) * contacts_limit
    
    contacts = []
    total_contacts = 0
    unique_tags = []
    broadcasts = []
    
    try:
        contacts = await db.list_contacts(
            bot_id,
            search=search_contacts,
            tag=tag_filter,
            limit=contacts_limit,
            offset=contacts_offset
        )
        total_contacts = await db.count_contacts(bot_id, search=search_contacts, tag=tag_filter)
        unique_tags = await db.list_contact_tags(bot_id)
        broadcasts = await db.list_broadcasts(bot_id, limit=50)
    except Exception as exc:
        log.error(f"Error fetching contacts/broadcasts state: {exc}")

    # Fetch Conversations and CRM info for client tabs
    await db.qualify_leads_with_action_link(
        config.QUALIFIED_CTA_URL,
        bot_id=bot_id,
    )
    
    threads = await db.list_conversation_threads(limit=100, bot_id=bot_id)
    selected_wa_id = wa_id or (threads[0]["wa_id"] if threads else None)
    messages = await db.list_conversation_messages(selected_wa_id, limit=120, bot_id=bot_id) if selected_wa_id else []
    selected_lead = await db.get_lead(selected_wa_id, bot_id=bot_id) if selected_wa_id else None
    
    # CRM info
    crm_counts = await db.crm_counts(bot_id=bot_id)
    crm_leads = await db.list_leads(
        None if status == "todos" else status,
        limit=200,
        bot_id=bot_id,
    )

    thread_items = "".join(
        f"""
        <a class="chat-item {'active' if selected_wa_id == t['wa_id'] else ''}" href="/client/app?bot_id={bot_id}&tab=conversations&wa_id={html.escape(t["wa_id"])}">
          <div class="chat-item-header">
            <span class="chat-item-name">{html.escape(_display_name(t.get("nombre"), t["wa_id"]))}</span>
            <span class="chat-item-time">{_fmt_dt(t.get("last_message_at")).split(" ")[-1]}</span>
          </div>
          <div class="chat-item-preview">{html.escape(_clip(t.get("last_content"), 60))}</div>
        </a>
        """
        for t in threads
    ) or '<div class="empty" style="padding: 20px; text-align: center; color: var(--muted);">Aún no hay conversaciones.</div>'

    bubble_html = "".join(
        f"""
        <div class="bubble {html.escape(m["role"])}">
          <div class="meta">{html.escape("Bot" if m["role"] == "assistant" else "Cliente")} - {_fmt_dt(m.get("created_at"))}</div>
          <div>{html.escape(m["content"])}</div>
        </div>
        """
        for m in messages
    ) or '<div class="empty" style="padding: 20px; text-align: center; color: var(--muted);">Selecciona una conversación para ver el historial.</div>'

    chat_header = ""
    crm_sidebar = ""
    if selected_wa_id:
        lead_status = (selected_lead or {}).get("qualification_status", "en_progreso")
        name_display = html.escape(_display_name((selected_lead or {}).get("nombre"), selected_wa_id))
        initials = name_display[:2].upper() if name_display else "??"
        
        chat_header = f"""
        <div class="chat-header">
          <div class="chat-avatar">{initials}</div>
          <div class="chat-header-info">
            <strong>{name_display}</strong>
            <small>{html.escape(selected_wa_id)}</small>
          </div>
        </div>
        """
        
        crm_sidebar = f"""
        <div class="chat-crm">
          <div class="crm-section">
            <h3>Detalles del Contacto</h3>
            <div style="font-weight:600;font-size:15px;margin-bottom:4px;">{name_display}</div>
            <div style="color:var(--muted);font-size:13px;margin-bottom:12px;">{html.escape((selected_lead or {}).get("negocio") or "Sin negocio")}</div>
            <a class="btn whatsapp" href="{_wa_link(selected_wa_id)}" target="_blank" style="width:100%; justify-content:center; display:flex; align-items:center; gap:8px; background:#25d366; color:white; border:none; text-decoration:none; border-radius:8px; padding:10px; font-weight:600; font-size:13px;">Conectar por WhatsApp</a>
          </div>
          <div class="crm-section">
            <h3>Estado del Lead</h3>
            <span class="badge b-{html.escape(lead_status)}">{html.escape(lead_status.replace("_", " "))}</span>
          </div>
          <div class="crm-section">
            <h3>Notas</h3>
            <div style="font-size:13px;color:var(--muted);background:#f8fafc;padding:10px;border-radius:8px;border:1px solid var(--line);">
              {html.escape((selected_lead or {}).get("notes") or "Sin notas guardadas.")}
            </div>
          </div>
        </div>
        """

    crm_tabs = [
        ("en_progreso", "No cualificados"),
        ("calificado", "Calificados"),
        ("descalificado", "Descalificados"),
        ("todos", "Todos"),
    ]
    crm_tabs_html = "".join(
        f'<a class="{"active" if key == status else ""}" href="/client/app?bot_id={bot_id}&tab=crm&status={key}">{label} ({sum(crm_counts.values()) if key == "todos" else crm_counts.get(key, 0)})</a>'
        for key, label in crm_tabs
    )
    
    crm_rows = "".join(
        f"""
        <tr>
          <td><strong>{html.escape(_display_name(l.get("nombre"), l["wa_id"]))}</strong><br><span class="muted">{html.escape(l["wa_id"])}</span></td>
          <td>{html.escape(l.get("negocio") or "-")}</td>
          <td><span class="badge b-{html.escape(l["qualification_status"])}">{html.escape(l["qualification_status"].replace("_", " "))}</span><br><span class="muted">{html.escape(l.get("disqualify_reason") or "")}</span></td>
          <td>{_fmt_dt(l.get("updated_at"))}</td>
          <td>
            <div class="actions" style="display:flex; gap:6px;">
              <form class="inline" method="post" action="/client/bots/{bot_id}/crm/{html.escape(l["wa_id"])}/status"><button class="btn secondary" name="status" value="en_progreso" style="padding:6px 10px; font-size:12px; cursor:pointer;">No cualificado</button></form>
              <form class="inline" method="post" action="/client/bots/{bot_id}/crm/{html.escape(l["wa_id"])}/status"><button class="btn primary-btn" name="status" value="calificado" style="padding:6px 10px; font-size:12px; cursor:pointer;">Calificar</button></form>
              <form class="inline" method="post" action="/client/bots/{bot_id}/crm/{html.escape(l["wa_id"])}/status"><button class="btn secondary" name="status" value="descalificado" style="padding:6px 10px; font-size:12px; color:var(--red); cursor:pointer;">Descartar</button></form>
            </div>
          </td>
        </tr>
        """
        for l in crm_leads
    ) or '<tr><td colspan="5" class="empty" style="text-align:center; padding:20px; color:var(--muted);">No hay leads en esta etapa.</td></tr>'

    bot_status = selected_bot.get("status") or "active"
    if bot_status == "active":
        bot_status_class = "success"
        bot_status_label = "Activo"
        bot_dot_color = "#10b981"
        bot_status_title = "Haz clic para PAUSAR el bot (modo humano únicamente)"
    else:
        bot_status_class = "warning"
        bot_status_label = "Pausado"
        bot_dot_color = "#f59e0b"
        bot_status_title = "Haz clic para REANUDAR las respuestas autónomas del bot"

    import base64
    # 2. TEMPLATE PANEL HTML
    templates_rows = ""
    for t in templates:
        t_status = (t.get("status") or "").upper()
        badge_class = "success" if t_status == "APPROVED" else ("danger" if t_status == "REJECTED" else "warning")
        tpl_status_label = "Aprobado" if t_status == "APPROVED" else ("Rechazado" if t_status == "REJECTED" else t.get("status") or "-")
        t_json = json.dumps(t)
        t_b64 = base64.b64encode(t_json.encode("utf-8")).decode("utf-8")
        templates_rows += f"""
        <tr>
          <td style="font-weight:600; font-size:13px;">
            <a href="#" class="template-link" onclick="event.preventDefault(); showTemplateDetailsB64('{t_b64}')" style="text-decoration:none; color:var(--primary); font-weight:600;">{html.escape(t.get("name") or "-")}</a>
            <br><span class="muted-text" style="font-size:11px;">{html.escape(t.get("language") or "-")}</span>
          </td>
          <td><span class="badge" style="font-size:11px;">{html.escape(t.get("category") or "-")}</span></td>
          <td><span class="badge {badge_class}" style="font-size:11px;">{html.escape(tpl_status_label)}</span></td>
        </tr>
        """
        
    if not templates_rows:
        templates_rows = '<tr><td colspan="3" style="text-align:center; padding:20px; color:var(--muted); font-size:13px;">Sin plantillas registradas en Meta para este número.</td></tr>'

    templates_table = f"""
    <div style="overflow-x:auto;">
      <table style="width:100%;">
        <thead>
          <tr>
            <th>Plantilla</th>
            <th>Categoría</th>
            <th>Estado</th>
          </tr>
        </thead>
        <tbody>
          {templates_rows}
        </tbody>
      </table>
    </div>
    """

    templates_error_html = f'<div class="notice-banner error" style="margin-bottom: 15px;">{ICONS["error"]} {html.escape(templates_error)}</div>' if templates_error else ""

    vars_regex = r"/\{\{\s*(\d+)\s*\}\}/g"
    amp_regex = r"/&/g"
    lt_regex = r"/</g"
    gt_regex = r"/>/g"
    bold_regex = r"/\*([^*]+)\*/g"
    italic_regex = r"/_([^_]+)_/g"
    strike_regex = r"/~([^~]+)~/g"
    nl_regex = r"/\n/g"

    templates_panel_html = f"""
    <div id="panel-templates" class="tab-panel">
      {templates_error_html}
      <div class="grid-2">
        <div class="card">
          <div class="card-header">
            <h2>Plantillas de Mensajes</h2>
            <p>Lista de plantillas aprobadas por Meta para iniciar conversaciones fuera de la ventana de 24 horas.</p>
          </div>
          {templates_table}
        </div>
        
        <div>
          <div class="card">
            <div class="card-header">
              <h2>Crear Plantilla en Meta</h2>
              <p>Envía una nueva plantilla para revisión y aprobación de Meta.</p>
            </div>
            <form method="post" action="/client/bots/{bot_id}/whatsapp/templates">
              <label>Nombre de la Plantilla</label>
              <input name="name" placeholder="bienvenida_cliente_v1" required style="margin-bottom:12px;" {"readonly" if session["role"] == "client_viewer" else ""}>
              
              <label>Idioma</label>
              <input name="language" value="es_MX" required style="margin-bottom:12px;" {"readonly" if session["role"] == "client_viewer" else ""}>
              
              <label>Categoría</label>
              <select name="category" style="margin-bottom:12px;" {"disabled" if session["role"] == "client_viewer" else ""}>
                <option value="UTILITY">Utility (Servicio, Seguimiento)</option>
                <option value="MARKETING">Marketing (Promociones, Ofertas)</option>
              </select>
              
              <label>Cuerpo de la Plantilla</label>
              <p class="muted-text" style="font-size:12px; margin-top:2px; margin-bottom:6px;">Usa <code>{{{{1}}}}</code>, <code>{{{{2}}}}</code> para agregar variables. Ejemplo: <i>Hola {{{{1}}}}, tu código es {{{{2}}}}</i>.</p>
              <textarea name="body_text" style="min-height: 100px; margin-bottom:12px;" placeholder="Hola {{{{1}}}}, tu cita está agendada para el {{{{2}}}}." required {"readonly" if session["role"] == "client_viewer" else ""}></textarea>
              
              <div id="clientTemplateVarsContainer" style="display:none; margin-bottom:16px;">
                <label style="font-weight:600; display:block; margin-top:12px;">Muestras de Variables</label>
                <p class="muted-text" style="font-size:11px; margin-top:2px; margin-bottom:8px;">Meta requiere un ejemplo real para aprobar cada variable de tu plantilla.</p>
                <div id="clientTemplateVarsInputs" style="display:flex; flex-direction:column; gap:8px;"></div>
              </div>
              
              <div style="margin-top:14px;">
                {"<button class='btn primary-btn' type='submit'>Enviar plantilla</button>" if session["role"] != "client_viewer" else "<span class='badge'>Solo lectura</span>"}
              </div>
            </form>
          </div>
          
          <!-- Live Preview Card -->
          <div class="card" style="margin-top:20px; border:1px solid #bae6fd; background:#f0f9ff; padding:15px; border-radius:8px;">
            <div class="card-header" style="padding-bottom:8px; border-bottom:1px solid #bae6fd; margin-bottom:10px;">
              <h3 style="margin:0; font-size:14px; font-weight:700; color:#0369a1; display:flex; align-items:center; gap:6px;">
                <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" style="width:16px; height:16px;"><path d="M2 12s3-7 10-7 10 7 10 7-3 7-10 7-10-7-10-7Z"/><circle cx="12" cy="12" r="3"/></svg>
                Vista Previa de WhatsApp (Real-time)
              </h3>
            </div>
            <div style="background:#e5ddd5; padding:12px; border-radius:6px; min-height:80px; position:relative; font-family:\'Inter\', sans-serif;">
              <div style="background:#ffffff; padding:8px 10px; border-radius:6px; max-width:90%; font-size:13px; line-height:1.4; color:#000000; box-shadow:0 1px 0.5px rgba(0,0,0,0.13); position:relative;">
                <span id="clientTemplatePreviewBody" style="word-break: break-word;">Escribe el texto de tu plantilla...</span>
                <div style="text-align:right; font-size:10px; color:#a0a0a0; margin-top:4px;">12:00 PM</div>
              </div>
            </div>
          </div>
        </div>
      </div>
      
      <script>
        (function() {{
          const bodyText = document.querySelector(\'#panel-templates textarea[name="body_text"]\');
          const varsContainer = document.getElementById(\'clientTemplateVarsContainer\');
          const varsInputs = document.getElementById(\'clientTemplateVarsInputs\');
          const previewBody = document.getElementById(\'clientTemplatePreviewBody\');

          function updateTemplatePreview() {{
            let text = bodyText.value || "Escribe el texto de tu plantilla...";
            
            const inputs = varsInputs.querySelectorAll(\'input[name="examples"]\');
            inputs.forEach(input => {{
              const varNum = input.dataset.varNum;
              const val = input.value.trim() || `{{{{${{varNum}}}}}}`;
              text = text.replaceAll(\'{{\' + varNum + \'}}\', val);
            }});

            let formattedText = text
              .replace({amp_regex}, "&amp;")
              .replace({lt_regex}, "&lt;")
              .replace({gt_regex}, "&gt;")
              .replace({bold_regex}, \'<strong>$1</strong>\')
              .replace({italic_regex}, \'<em>$1</em>\')
              .replace({strike_regex}, \'<del>$1</del>\')
              .replace({nl_regex}, \'<br>\');

            previewBody.innerHTML = formattedText;
          }}

          if (bodyText) {{
            function detectVariables() {{
              const text = bodyText.value;
              const regex = {vars_regex};
              let match;
              const detectedVars = new Set();
              while ((match = regex.exec(text)) !== null) {{
                detectedVars.add(parseInt(match[1]));
              }}

              const sortedVars = Array.from(detectedVars).sort((a, b) => a - b);

              const existingValues = {{}};
              varsInputs.querySelectorAll(\'input[name="examples"]\').forEach(input => {{
                existingValues[input.dataset.varNum] = input.value;
              }});

              varsInputs.innerHTML = "";
              if (sortedVars.length > 0) {{
                varsContainer.style.display = "block";
                sortedVars.forEach(varNum => {{
                  const row = document.createElement(\'div\');
                  row.style.cssText = \'display:flex; flex-direction:column; gap:4px; margin-bottom:8px;\';
                  
                  const label = document.createElement(\'label\');
                  label.style.cssText = \'margin:0; font-size:12px; font-weight:600; color:var(--muted);\';
                  label.textContent = `Muestra para {{{{${{varNum}}}}}}`;

                  const input = document.createElement(\'input\');
                  input.type = \'text\';
                  input.name = \'examples\';
                  input.placeholder = `Ej. para {{{{${{varNum}}}}}}`;
                  input.dataset.varNum = varNum;
                  input.required = true;
                  input.value = existingValues[varNum] || "";
                  input.style.margin = \'0\';
                  
                  input.addEventListener(\'input\', updateTemplatePreview);
                  
                  row.appendChild(label);
                  row.appendChild(input);
                  varsInputs.appendChild(row);
                }});
              }} else {{
                varsContainer.style.display = "none";
              }}

              updateTemplatePreview();
            }}

            bodyText.addEventListener(\'input\', detectVariables);
            // Run on load to pick up any prefilled values
            detectVariables();
          }}
        }})();
      </script>
    </div>
    """
    # 2b. CONTACTS PANEL HTML
    contacts_rows = ""
    for c in contacts:
        c_tags = html.escape(c.get("tags") or "")
        tags_badges = "".join(f'<span class="badge" style="margin-right:4px; font-size:10.5px; background:#f1f5f9; color:#475569; border:1px solid #e2e8f0; border-radius:4px; padding:2px 6px;">{t.strip()}</span>' for t in c_tags.split(",") if t.strip())
        contacts_rows += f"""
        <tr>
          <td style="width: 40px; text-align: center;"><input type="checkbox" name="selected_contacts" value="{html.escape(c['wa_id'])}" class="contact-checkbox" onchange="updateSelectedCount()"></td>
          <td style="font-weight:600; font-size:13px; color:var(--ink);">{html.escape(c.get("name") or "-")}</td>
          <td style="font-family:monospace; font-size:12.5px; color:var(--ink);">{html.escape(c.get("wa_id") or "-")}</td>
          <td>{html.escape(c.get("business") or "-")}</td>
          <td>{tags_badges}</td>
          <td style="font-size:11.5px; color:var(--muted);">{_fmt_dt(c.get("created_at"))}</td>
        </tr>
        """
        
    if not contacts_rows:
        contacts_rows = '<tr><td colspan="6" style="text-align:center; padding:30px; color:var(--muted); font-size:13.5px;">No se encontraron contactos en el directorio.</td></tr>'

    tag_options = '<option value="">Todas las etiquetas</option>'
    for t in unique_tags:
        selected_attr = "selected" if t == tag_filter else ""
        tag_options += f'<option value="{html.escape(t)}" {selected_attr}>{html.escape(t)}</option>'

    total_pages = max(1, (total_contacts + contacts_limit - 1) // contacts_limit)
    prev_disabled = "disabled" if cpage <= 1 else ""
    next_disabled = "disabled" if cpage >= total_pages else ""

    # Check if we are in Column Mapping wizard
    if map_file:
        tmp_dir = os.path.join(os.getcwd(), "tmp", "uploads")
        file_path = os.path.join(tmp_dir, map_file)
        headers = extract_headers_from_file(file_path)
        
        if not headers:
            mapping_wizard_html = f"""
            <div class="card" style="margin-top:20px; border:1px solid var(--red); background:#fff1f2;">
              <div class="card-header">
                <h2 style="color:var(--red);">{ICONS["error"]} Error de Lectura</h2>
                <p>No pudimos extraer columnas de tu archivo. Asegúrate de que tenga una fila de encabezados y no esté dañado.</p>
              </div>
              <a href="/client/app?bot_id={bot_id}&tab=contacts" class="btn secondary">Volver al directorio</a>
            </div>
            """
        else:
            col_options = ""
            for idx, h in enumerate(headers):
                col_options += f'<option value="{idx}">{html.escape(h)} (Columna {idx+1})</option>'
                
            mapping_wizard_html = f"""
            <div class="card" style="margin-top:20px; border:1px solid #bae6fd; background:#f0f9ff;">
              <div class="card-header">
                <h2 style="color:#0369a1;">Mapear Columnas de Importación</h2>
                <p>Asocia las columnas de tu archivo <code>{html.escape(map_file.split(".",1)[-1] or "cargado")}</code> con los atributos de contactos en Asistto.</p>
              </div>
              <form method="post" action="/client/bots/{bot_id}/contacts/import">
                <input type="hidden" name="file_token" value="{html.escape(map_file)}">
                
                <div style="display:flex; flex-direction:column; gap:14px; max-width:500px;">
                  <div>
                    <label style="font-weight:600; margin-bottom:4px;">Columna de Teléfono (Obligatoria)</label>
                    <p class="muted-text" style="font-size:11px; margin:0 0 6px 0;">Debe contener los números telefónicos con prefijo de país.</p>
                    <select name="phone_col" required>
                      <option value="" disabled selected>-- Seleccionar columna --</option>
                      {col_options}
                    </select>
                  </div>
                  
                  <div>
                    <label style="font-weight:600; margin-bottom:4px;">Columna de Nombre (Opcional)</label>
                    <select name="name_col">
                      <option value="-1">-- No importar (vacío) --</option>
                      {col_options}
                    </select>
                  </div>
                  
                  <div>
                    <label style="font-weight:600; margin-bottom:4px;">Columna de Negocio/Empresa (Opcional)</label>
                    <select name="business_col">
                      <option value="-1">-- No importar (vacío) --</option>
                      {col_options}
                    </select>
                  </div>
                  
                  <div>
                    <label style="font-weight:600; margin-bottom:4px;">Columna de Etiquetas (Opcional)</label>
                    <p class="muted-text" style="font-size:11px; margin:0 0 6px 0;">Las etiquetas dentro de la celda deben estar separadas por comas.</p>
                    <select name="tags_col">
                      <option value="-1">-- No importar (vacío) --</option>
                      {col_options}
                    </select>
                  </div>
                </div>
                
                <div style="margin-top:24px; display:flex; gap:10px;">
                  <button class="btn primary-btn" type="submit">Completar Importación</button>
                  <a href="/client/app?bot_id={bot_id}&tab=contacts" class="btn secondary" style="text-decoration:none; display:inline-flex; align-items:center;">Cancelar</a>
                </div>
              </form>
            </div>
            """
        
        contacts_panel_html = f"""
        <div id="panel-contacts" class="tab-panel">
          {mapping_wizard_html}
        </div>
        """
    else:
        contacts_panel_html = f"""
        <div id="panel-contacts" class="tab-panel">
          <div class="grid-contacts" style="display:grid; grid-template-columns:1fr 340px; gap:20px; align-items:start;">
            
            <!-- Left Side: Table -->
            <div class="card">
              <div class="card-header" style="display:flex; justify-content:space-between; align-items:center; flex-wrap:wrap; gap:10px;">
                <div>
                  <h2>Directorio de Contactos</h2>
                  <p>Catálogo de contactos para campañas de difusión por WhatsApp. Total: <strong>{total_contacts}</strong> contactos.</p>
                </div>
              </div>
              
              <!-- Filter & Search Bar -->
              <div style="background:#f8fafc; padding:12px 16px; border-bottom:1px solid var(--line); display:flex; gap:10px; flex-wrap:wrap; align-items:center;">
                <form method="get" action="/client/app" style="display:flex; gap:10px; flex:1; margin:0; flex-wrap:wrap;">
                  <input type="hidden" name="bot_id" value="{bot_id}">
                  <input type="hidden" name="tab" value="contacts">
                  <input name="search_q" placeholder="Buscar por nombre, teléfono..." value="{html.escape(search_q or '')}" style="max-width:240px; margin:0; font-size:13px; padding:6px 10px;">
                  <select name="tag_q" style="max-width:180px; margin:0; font-size:13px; padding:6px 10px;">
                    {tag_options}
                  </select>
                  <button class="btn" type="submit" style="padding:6px 14px; font-size:13px;">Filtrar</button>
                  <a href="/client/app?bot_id={bot_id}&tab=contacts" class="btn secondary" style="text-decoration:none; display:inline-flex; align-items:center; padding:6px 14px; font-size:13px;">Limpiar</a>
                </form>
                
                <div id="contactsActionBar" style="display:none; gap:8px;">
                  <button class="btn" style="background:#2563eb; color:white; border:none; padding:6px 12px; font-size:13px;" onclick="startBroadcastFromSelected()">Crear Difusión ({html.escape("{{count}}")})</button>
                  <form method="post" action="/client/bots/{bot_id}/contacts/delete" id="deleteContactsForm" style="margin:0; padding:0; display:inline;">
                    <input type="hidden" name="selected_contacts_list" id="selectedContactsListInput">
                    <button class="btn secondary" type="button" style="color:var(--red); padding:6px 12px; font-size:13px;" onclick="confirmDeleteContacts()">Eliminar</button>
                  </form>
                </div>
              </div>

              <div style="overflow-x:auto;">
                <table style="width:100%;" id="contactsTable">
                  <thead>
                    <tr>
                      <th style="width: 40px; text-align: center;"><input type="checkbox" id="selectAllContacts" onchange="toggleSelectAllContacts(this)"></th>
                      <th>Nombre</th>
                      <th>Teléfono (WA ID)</th>
                      <th>Negocio</th>
                      <th>Etiquetas</th>
                      <th>Registro</th>
                    </tr>
                  </thead>
                  <tbody>
                    {contacts_rows}
                  </tbody>
                </table>
              </div>
              
              <!-- Paging footer -->
              <div style="padding:14px; border-top:1px solid var(--line); display:flex; justify-content:space-between; align-items:center;">
                <span style="font-size:12.5px; color:var(--muted);">Pág. {cpage} de {total_pages}</span>
                <div style="display:flex; gap:6px;">
                  <a href="/client/app?bot_id={bot_id}&tab=contacts&search_q={html.escape(search_q or '')}&tag_q={html.escape(tag_q or '')}&cpage={cpage - 1}" class="btn secondary {'disabled' if cpage <= 1 else ''}" style="padding:5px 12px; font-size:12px; text-decoration:none;">Anterior</a>
                  <a href="/client/app?bot_id={bot_id}&tab=contacts&search_q={html.escape(search_q or '')}&tag_q={html.escape(tag_q or '')}&cpage={cpage + 1}" class="btn secondary {'disabled' if cpage >= total_pages else ''}" style="padding:5px 12px; font-size:12px; text-decoration:none;">Siguiente</a>
                </div>
              </div>
            </div>
            
            <!-- Right Side: Import & Manual creation -->
            <div style="display:flex; flex-direction:column; gap:20px;">
              <div class="card">
                <div class="card-header">
                  <h2>Importar Contactos</h2>
                  <p>Carga un archivo de Excel (.xlsx) o texto (.csv) con tu lista de contactos para importarlos masivamente.</p>
                </div>
                <form method="post" action="/client/bots/{bot_id}/contacts/upload" enctype="multipart/form-data">
                  <label>Seleccionar Archivo</label>
                  <input type="file" name="file" accept=".csv,.xlsx" required style="font-size:13px; color:var(--muted); margin-bottom:12px;">
                  <button class="btn primary-btn" type="submit" {"disabled" if session["role"] == "client_viewer" else ""}>Subir y Mapear</button>
                </form>
              </div>
              
              <div class="card">
                <div class="card-header">
                  <h2>Agregar Contacto Nuevo</h2>
                  <p>Registra un contacto de forma individual en tu directorio.</p>
                </div>
                <form method="post" action="/client/bots/{bot_id}/contacts/create-manual">
                  <label>Nombre Completo</label>
                  <input name="name" placeholder="Ej. Juan Pérez" required style="margin-bottom:10px;" {"readonly" if session["role"] == "client_viewer" else ""}>
                  
                  <label>Teléfono (WhatsApp con prefijo)</label>
                  <input name="wa_id" placeholder="Ej. 5216869032840" required style="margin-bottom:10px;" {"readonly" if session["role"] == "client_viewer" else ""}>
                  
                  <label>Nombre del Negocio</label>
                  <input name="business" placeholder="Ej. Tacos El Pastor" style="margin-bottom:10px;" {"readonly" if session["role"] == "client_viewer" else ""}>
                  
                  <label>Etiquetas (Separadas por coma)</label>
                  <input name="tags" placeholder="Ej. VIP, Prospecto" style="margin-bottom:14px;" {"readonly" if session["role"] == "client_viewer" else ""}>
                  
                  <button class="btn primary-btn" type="submit" {"disabled" if session["role"] == "client_viewer" else ""}>Registrar Contacto</button>
                </form>
              </div>
            </div>
            
          </div>
          
          <script>
            function toggleSelectAllContacts(master) {{
              const checkboxes = document.querySelectorAll('.contact-checkbox');
              checkboxes.forEach(cb => cb.checked = master.checked);
              updateSelectedCount();
            }}
            
            function updateSelectedCount() {{
              const selected = document.querySelectorAll('.contact-checkbox:checked');
              const actionBar = document.getElementById('contactsActionBar');
              if (selected.length > 0) {{
                actionBar.style.display = 'flex';
                const btn = actionBar.querySelector('button');
                btn.textContent = `Crear Difusión (${{selected.length}})`;
              }} else {{
                actionBar.style.display = 'none';
              }}
            }}
            
            function startBroadcastFromSelected() {{
              const selected = document.querySelectorAll('.contact-checkbox:checked');
              const ids = Array.from(selected).map(cb => cb.value);
              const url = new URL(window.location.href);
              url.searchParams.set("tab", "campaigns");
              url.searchParams.set("broadcast_selected", ids.join(","));
              window.location.href = url.toString();
            }}
            
            function confirmDeleteContacts() {{
              if (confirm("¿Estás seguro de que deseas eliminar los contactos seleccionados? Esta acción es irreversible.")) {{
                const selected = document.querySelectorAll('.contact-checkbox:checked');
                const ids = Array.from(selected).map(cb => cb.value);
                const form = document.getElementById('deleteContactsForm');
                
                ids.forEach(id => {{
                  const input = document.createElement('input');
                  input.type = 'hidden';
                  input.name = 'selected_contacts';
                  input.value = id;
                  form.appendChild(input);
                }});
                
                form.submit();
              }}
            }}
          </script>
        </div>
        """

    # 2c. CAMPAIGNS PANEL HTML
    broadcast_selected_count = 0
    selected_wa_ids_str = ""
    if broadcast_selected:
        selected_ids_list = [w.strip() for w in broadcast_selected.split(",") if w.strip()]
        broadcast_selected_count = len(selected_ids_list)
        selected_wa_ids_str = broadcast_selected

    broadcast_rows = ""
    for b in broadcasts:
        b_status = b["status"].upper()
        badge_class = "success" if b_status == "COMPLETED" else ("warning" if b_status == "RUNNING" else ("danger" if b_status == "FAILED" else "info"))
        b_status_label = "Completada" if b_status == "COMPLETED" else ("Enviando..." if b_status == "RUNNING" else ("Fallida" if b_status == "FAILED" else b["status"]))
        
        total = b["total_recipients"] or 0
        sent = b["sent_count"] or 0
        failed = b["failed_count"] or 0
        progress_pct = int(min(100, (sent + failed) * 100 / total)) if total > 0 else 0
        
        broadcast_rows += f"""
        <tr>
          <td style="font-weight:600; font-size:13px; color:var(--ink);">{html.escape(b["name"])}</td>
          <td><code>{html.escape(b["template_name"])}</code></td>
          <td style="font-size:12px;"><strong>{sent + failed}</strong> / {total} <span class="muted-text">({progress_pct}%)</span><br>
            <div style="width:100%; background:#e2e8f0; height:6px; border-radius:3px; overflow:hidden; margin-top:4px;">
              <div style="width:{progress_pct}%; background:var(--primary); height:100%;"></div>
            </div>
            <small style="color:var(--green); font-weight:600;">{sent} exitosos</small> | <small style="color:var(--red); font-weight:600;">{failed} fallidos</small>
          </td>
          <td><span class="badge {badge_class}" style="font-size:11px;">{html.escape(b_status_label)}</span></td>
          <td style="font-size:11.5px; color:var(--muted);">{_fmt_dt(b.get("created_at"))}</td>
        </tr>
        """
        
    if not broadcast_rows:
        broadcast_rows = '<tr><td colspan="5" style="text-align:center; padding:30px; color:var(--muted); font-size:13.5px;">No se han realizado campañas de envío masivo.</td></tr>'

    template_options = '<option value="" disabled selected>-- Seleccionar plantilla --</option>'
    for t in templates:
        if (t.get("status") or "").upper() == "APPROVED":
            template_options += f'<option value="{html.escape(t.get("name"))}">{html.escape(t.get("name"))}</option>'

    templates_json = json.dumps(templates)

    create_view_display = "block" if broadcast_selected else "none"
    list_view_display = "none" if broadcast_selected else "block"

    campaigns_panel_html = f"""
    <div id="panel-campaigns" class="tab-panel">
      
      <!-- List Campaigns View -->
      <div id="campaignsListView" style="display:{list_view_display};">
        <div class="card">
          <div class="card-header" style="display:flex; justify-content:space-between; align-items:center; flex-wrap:wrap; gap:10px;">
            <div>
              <h2>Campañas de Envío Masivo</h2>
              <p>Historial y estado de las campañas de difusión por WhatsApp usando plantillas aprobadas.</p>
            </div>
            {"<button class='btn primary-btn' onclick='showNewCampaignForm()'>Nueva Campaña</button>" if session["role"] != "client_viewer" else ""}
          </div>
          
          <div style="overflow-x:auto;">
            <table style="width:100%;">
              <thead>
                <tr>
                  <th>Nombre de Campaña</th>
                  <th>Plantilla Utilizada</th>
                  <th>Progreso de Envío</th>
                  <th>Estado</th>
                  <th>Fecha de Creación</th>
                </tr>
              </thead>
              <tbody>
                {broadcast_rows}
              </tbody>
            </table>
          </div>
        </div>
      </div>
      
      <!-- Create Campaign View -->
      <div id="campaignsCreateView" style="display:{create_view_display};">
        <div class="grid-2">
          
          <!-- Left side: Configuration -->
          <div class="card">
            <div class="card-header">
              <h2>Configurar Nueva Campaña</h2>
              <p>Elige tu plantilla y configura el mapeo de variables dinámicas.</p>
            </div>
            
            <form method="post" action="/client/bots/{bot_id}/campaigns/create">
              <input type="hidden" name="language_code" id="campaignLangCode" value="es_MX">
              <input type="hidden" name="vars_count" id="campaignVarsCountInput" value="0">
              
              <label>Nombre de la Campaña</label>
              <input name="campaign_name" placeholder="Ej. Promoción Junio 2026" required style="margin-bottom:12px;">
              
              <label>Destinatarios de la Campaña</label>
              <div style="margin-bottom:12px; display:flex; flex-direction:column; gap:8px;">
                <div style="display:flex; gap:16px;">
                  <label style="display:inline-flex; align-items:center; gap:6px; font-weight:normal; margin:0; cursor:pointer;">
                    <input type="radio" name="recipients_option" value="selected" {"checked" if broadcast_selected else "disabled"} onchange="toggleRecipientsType(this.value)"> 
                    Contactos Seleccionados ({broadcast_selected_count})
                  </label>
                  <label style="display:inline-flex; align-items:center; gap:6px; font-weight:normal; margin:0; cursor:pointer;">
                    <input type="radio" name="recipients_option" value="all" {"" if broadcast_selected else "checked"} onchange="toggleRecipientsType(this.value)"> 
                    Todos los Contactos ({total_contacts})
                  </label>
                </div>
                <input type="hidden" name="selected_wa_ids" value="{html.escape(selected_wa_ids_str)}">
              </div>
              
              <label>Seleccionar Plantilla Meta</label>
              <select name="template_name" required style="margin-bottom:16px;" onchange="onTemplateSelected(this)">
                {template_options}
              </select>
              
              <!-- Dynamic Variables Container -->
              <div id="campaignVarsContainer" style="display:none; margin-bottom:20px;">
                <label style="font-weight:700; margin-bottom:8px; display:block;">Mapear Variables Dinámicas</label>
                <p class="muted-text" style="font-size:11.5px; margin-top:0; margin-bottom:10px;">Asocia cada marcador <code>{{{{1}}}}</code>, <code>{{{{2}}}}</code> con campos del contacto.</p>
                <div id="campaignVarsInputs"></div>
              </div>
              <label>Confirmacion de envio</label>
              <input name="confirm_send" placeholder="Escribe CONFIRMAR para iniciar" required style="margin-bottom:16px;">
              
              <div style="display:flex; gap:10px;">
                <button class="btn primary-btn" type="submit">Iniciar Campaña Masiva</button>
                <button class="btn secondary" type="button" onclick="cancelNewCampaign()">Cancelar</button>
              </div>
            </form>
          </div>
          
          <!-- Right side: Template details / Preview -->
          <div class="card" style="border:1px solid #bae6fd; background:#f0f9ff;">
            <div class="card-header" style="border-bottom:1px solid #bae6fd; padding-bottom:8px; margin-bottom:12px;">
              <h3 style="margin:0; font-size:14px; font-weight:700; color:#0369a1; display:flex; align-items:center; gap:6px;">
                <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" style="width:16px; height:16px;"><path d="M2 12s3-7 10-7 10 7 10 7-3 7-10 7-10-7-10-7Z"/><circle cx="12" cy="12" r="3"/></svg>
                Vista de Plantilla Seleccionada
              </h3>
            </div>
            <div style="background:#e5ddd5; padding:16px; border-radius:6px; min-height:100px; position:relative; font-family:\'Inter\', sans-serif;">
              <div style="background:#ffffff; padding:10px 12px; border-radius:6px; max-width:90%; font-size:13px; line-height:1.4; color:#000000; box-shadow:0 1px 0.5px rgba(0,0,0,0.13); position:relative;">
                <span id="campaignTemplateText" style="word-break: break-word; color:#333;">Selecciona una plantilla para previsualizar el cuerpo...</span>
                <div style="text-align:right; font-size:10px; color:#a0a0a0; margin-top:4px;">12:00 PM</div>
              </div>
            </div>
            <div style="margin-top:14px; font-size:12px; color:#0369a1;">
              <strong>Nota:</strong> Los destinatarios recibirán el mensaje reemplazando cada variable con sus datos correspondientes. Si la variable está mapeada al Nombre, se reemplazará por su nombre real guardado en el directorio.
            </div>
          </div>
          
        </div>
      </div>
      
      <script>
        const metaTemplates = {templates_json};
        
        function showNewCampaignForm() {{
          document.getElementById('campaignsListView').style.display = 'none';
          document.getElementById('campaignsCreateView').style.display = 'block';
        }}
        
        function cancelNewCampaign() {{
          document.getElementById('campaignsListView').style.display = 'block';
          document.getElementById('campaignsCreateView').style.display = 'none';
          const url = new URL(window.location.href);
          if (url.searchParams.has("broadcast_selected")) {{
            url.searchParams.delete("broadcast_selected");
            window.location.href = url.toString();
          }}
        }}
        
        function toggleRecipientsType(val) {{
          // logic placeholder
        }}
        
        function onTemplateSelected(select) {{
          const templateName = select.value;
          const selectedTpl = metaTemplates.find(t => t.name === templateName);
          const variablesContainer = document.getElementById('campaignVarsContainer');
          const variablesInputs = document.getElementById('campaignVarsInputs');
          const previewText = document.getElementById('campaignTemplateText');
          
          if (!selectedTpl) return;
          
          document.getElementById('campaignLangCode').value = selectedTpl.language || "es_MX";
          
          const bodyComp = selectedTpl.components.find(c => c.type === 'BODY');
          const bodyText = bodyComp ? bodyComp.text : "";
          previewText.innerHTML = bodyText.replace(/\\n/g, '<br>');
          
          const regex = /\\{{\\s*(\\d+)\\s*\\}}/g;
          let match;
          const vars = new Set();
          while ((match = regex.exec(bodyText)) !== null) {{
            vars.add(parseInt(match[1]));
          }}
          
          const sortedVars = Array.from(vars).sort((a, b) => a - b);
          document.getElementById('campaignVarsCountInput').value = sortedVars.length;
          
          variablesInputs.innerHTML = "";
          if (sortedVars.length > 0) {{
            variablesContainer.style.display = "block";
            sortedVars.forEach(varNum => {{
              const div = document.createElement('div');
              div.style.cssText = 'background:#f8fafc; border:1px solid #e2e8f0; border-radius:6px; padding:12px; margin-bottom:12px;';
              div.innerHTML = `
                <label style="font-weight:600; margin-bottom:4px; display:block; font-size:12.5px;">Variable {{{{${{varNum}}}}}}</label>
                <div style="display:flex; gap:10px; flex-wrap:wrap; align-items:center;">
                  <select name="var_map_type_${{varNum}}" style="margin:0; flex:1; max-width:200px;" onchange="toggleVarValueInput(${{varNum}}, this.value)">
                    <option value="name">Nombre del Contacto</option>
                    <option value="business">Negocio del Contacto</option>
                    <option value="wa_id">Teléfono del Contacto</option>
                    <option value="fixed">Valor Fijo (Texto estático)</option>
                  </select>
                  <input type="text" name="var_map_value_${{varNum}}" id="var_map_value_input_${{varNum}}" placeholder="Ej. Clinica Dental" style="display:none; margin:0; flex:2;" required disabled>
                </div>
              `;
              variablesInputs.appendChild(div);
            }});
          }} else {{
            variablesContainer.style.display = "none";
          }}
        }}
        
        function toggleVarValueInput(varNum, val) {{
          const input = document.getElementById(`var_map_value_input_${{varNum}}`);
          if (val === 'fixed') {{
            input.style.display = 'block';
            input.required = true;
            input.disabled = false;
          }} else {{
            input.style.display = 'none';
            input.required = false;
            input.disabled = true;
          }}
        }}
      </script>

      <!-- Modal para ver detalle de la plantilla -->
      <div id="templateDetailsModal" class="modal-overlay" onclick="closeTemplateDetailsModal(event)">
        <div class="modal-content" onclick="event.stopPropagation()">
          <div class="modal-header">
            <h3 id="modalTemplateName">Detalles de la Plantilla</h3>
            <button class="modal-close-btn" onclick="closeTemplateDetailsModal(event)">&times;</button>
          </div>
          <div class="modal-body" style="max-height: 70vh; overflow-y: auto;">
            <div style="display:grid; grid-template-columns: 1fr 1fr; gap:16px; margin-bottom: 20px;">
              <div>
                <span class="muted-text" style="font-size:11px; display:block; text-transform:uppercase; font-weight:600; color:var(--muted);">Idioma</span>
                <strong id="modalTemplateLang" style="font-size:14px; color:var(--ink);">-</strong>
              </div>
              <div>
                <span class="muted-text" style="font-size:11px; display:block; text-transform:uppercase; font-weight:600; color:var(--muted);">Categoría</span>
                <span id="modalTemplateCat" class="badge" style="font-size:11px; margin-top:2px;">-</span>
              </div>
              <div>
                <span class="muted-text" style="font-size:11px; display:block; text-transform:uppercase; font-weight:600; color:var(--muted);">Estado en Meta</span>
                <span id="modalTemplateStatus" class="badge" style="font-size:11px; margin-top:2px;">-</span>
              </div>
            </div>
            
            <div style="border: 1px solid var(--line); border-radius: 8px; padding: 16px; background:#f8fafc; margin-bottom: 20px;">
              <span class="muted-text" style="font-size:11px; display:block; text-transform:uppercase; font-weight:600; margin-bottom:8px; color:var(--muted);">Estructura de la Plantilla</span>
              <div id="modalTemplatePreview" style="font-size:13px; line-height:1.5; color:var(--ink);">
                Estructura de la plantilla...
              </div>
            </div>
            
            <div style="background: #fffbeb; border: 1px solid #fef3c7; border-radius: 8px; padding: 12px 16px; display: flex; gap: 10px; align-items: flex-start;">
              <svg viewBox="0 0 24 24" fill="none" stroke="#d97706" stroke-width="2" stroke-linecap="round" stroke-linejoin="round" style="width: 20px; height: 20px; flex-shrink:0; margin-top: 2px;">
                <circle cx="12" cy="12" r="10"></circle>
                <line x1="12" y1="16" x2="12" y2="12"></line>
                <line x1="12" y1="8" x2="12.01" y2="8"></line>
              </svg>
              <div style="font-size: 12px; color: #92400e; line-height: 1.4;">
                <strong>Nota sobre la edición:</strong> De acuerdo con las políticas de Meta, las plantillas aprobadas o en revisión no se pueden modificar directamente para no interrumpir campañas activas. Si necesitas hacer cambios, por favor crea una nueva plantilla en el panel lateral (ej. agregando un sufijo como <code>_v2</code> al nombre).
              </div>
            </div>
          </div>
          <div class="modal-footer">
            <button class="btn secondary" onclick="closeTemplateDetailsModal(event)" style="padding: 6px 14px; font-size:13px;">Cerrar</button>
          </div>
        </div>
      </div>

      <script>
        function showTemplateDetailsB64(b64Str) {{
          try {{
            const t = JSON.parse(atob(b64Str));
            document.getElementById('modalTemplateName').innerText = t.name || 'Detalles';
            document.getElementById('modalTemplateLang').innerText = t.language || '-';
            
            const catBadge = document.getElementById('modalTemplateCat');
            catBadge.innerText = t.category || '-';
            
            const statusBadge = document.getElementById('modalTemplateStatus');
            const status = (t.status || '').toUpperCase();
            statusBadge.innerText = status === 'APPROVED' ? 'Aprobado' : (status === 'REJECTED' ? 'Rechazado' : t.status || '-');
            
            statusBadge.className = 'badge';
            if (status === 'APPROVED') statusBadge.classList.add('success');
            else if (status === 'REJECTED') statusBadge.classList.add('danger');
            else statusBadge.classList.add('warning');
            
            const previewEl = document.getElementById('modalTemplatePreview');
            previewEl.innerHTML = '';
            
            if (t.components && t.components.length > 0) {{
              t.components.forEach(c => {{
                const type = (c.type || '').toUpperCase();
                const div = document.createElement('div');
                div.style.marginBottom = '12px';
                
                let label = '';
                let content = c.text || '';
                
                if (type === 'HEADER') {{
                  label = 'Encabezado';
                }} else if (type === 'BODY') {{
                  label = 'Cuerpo';
                }} else if (type === 'FOOTER') {{
                  label = 'Pie de página';
                }} else if (type === 'BUTTONS') {{
                  label = 'Botones';
                  content = JSON.stringify(c.buttons, null, 2);
                }}
                
                if (label) {{
                  div.innerHTML = `<span style="font-size:10px; font-weight:700; color:var(--muted); text-transform:uppercase; display:block; margin-bottom:2px;">${{label}}</span>` + 
                                  `<div style="background:white; border:1px solid var(--line); border-radius:6px; padding:10px; white-space:pre-wrap; font-family:monospace; font-size:12px;">${{content}}</div>`;
                  previewEl.appendChild(div);
                }}
              }});
            }} else {{
              previewEl.innerText = 'Sin componentes de texto definidos.';
            }}
            
            const modal = document.getElementById('templateDetailsModal');
            modal.classList.add('active');
          }} catch (e) {{
            console.error(e);
            alert('Error al mostrar los detalles de la plantilla.');
          }}
        }}
        
        function closeTemplateDetailsModal(event) {{
          if (event) event.preventDefault();
          const modal = document.getElementById('templateDetailsModal');
          modal.classList.remove('active');
        }}
      </script>
    </div>
    """

    # 3. RENDER SECTIONS
    body_html = f"""
    <div class="header">
      <div>
        <h1>{html.escape(selected_bot["name"])}</h1>
        <p>Configuración del asistente inteligente para tu negocio</p>
      </div>
      <div class="header-actions">
        <span class="badge" style="background: #e0f2fe; color: #0369a1; border: 1px solid #bae6fd;">ID del Bot: {bot_id}</span>
        <form action="/client/bots/{bot_id}/toggle-status" method="post" class="inline" style="margin:0; padding:0;">
          <button type="submit" class="badge {bot_status_class} badge-toggle" style="border: none; cursor: pointer; display: inline-flex; gap: 6px; align-items: center;" title="{bot_status_title}">
            <span style="display:inline-block; width:8px; height:8px; border-radius:50%; background:{bot_dot_color};"></span>
            Bot: {bot_status_label}
          </button>
        </form>
        <span class="badge">WhatsApp: {html.escape(wa_info.get("display_phone_number") or "sin conectar")}</span>
      </div>
    </div>
    
    <!-- 1. TAF PANEL: INICIO -->
    <div id="panel-inicio" class="tab-panel active">
      <div class="grid-3" style="margin-bottom: 24px;">
        <div class="kpi-card">
          <div class="kpi-title">Chats Activos</div>
          <div class="kpi-value">{metrics.get("conversations", 0)}</div>
          <div class="kpi-status ok">Historial guardado</div>
        </div>
        <div class="kpi-card">
          <div class="kpi-title">Mensajes Procesados</div>
          <div class="kpi-value">{metrics.get("messages", 0)}</div>
          <div class="kpi-status ok">Respuestas autónomas</div>
        </div>
        <div class="kpi-card">
          <div class="kpi-title">Leads Registrados</div>
          <div class="kpi-value">{metrics.get("leads", 0)}</div>
          <div class="kpi-status ok">CRM de Leads</div>
        </div>
      </div>
      
      <div class="grid-cards">
        <div>
          <div class="card">
            <div class="card-header">
              <h2>Conversaciones Recientes</h2>
              <p>Últimos hilos de chat iniciados en WhatsApp</p>
            </div>
            {_render_conversations_table(recent_threads)}
          </div>
        </div>
        <div>
          <div class="card">
            <div class="card-header">
              <h2>Estado del Asistente</h2>
              <p>Estado operacional de tus servicios</p>
            </div>
            <div style="display:flex; flex-direction:column; gap:12px; margin-top:14px;">
              <div style="display:flex; justify-content:space-between; align-items:center; padding-bottom:8px; border-bottom:1px solid var(--line);">
                <span class="bold-text" style="font-size:13px;">WhatsApp Cloud API</span>
                <span class="badge {"success" if wa_info.get("phone_number_id") else "danger"}">
                  {"Conectado" if wa_info.get("phone_number_id") else "Desconectado"}
                </span>
              </div>
              <div style="display:flex; justify-content:space-between; align-items:center; padding-bottom:8px; border-bottom:1px solid var(--line);">
                <span class="bold-text" style="font-size:13px;">Google Calendar</span>
                <span class="badge {"success" if calendar_status.get("enabled") else "warning"}">
                  {"Activo" if calendar_status.get("enabled") else "Inactivo"}
                </span>
              </div>
              <div style="display:flex; justify-content:space-between; align-items:center; padding-bottom:8px; border-bottom:1px solid var(--line);">
                <span class="bold-text" style="font-size:13px;">Reglas de Escalado</span>
                <span class="badge {"success" if escalate_enabled else "warning"}">
                  {"Habilitado" if escalate_enabled else "Deshabilitado"}
                </span>
              </div>
              <div style="display:flex; justify-content:space-between; align-items:center;">
                <span class="bold-text" style="font-size:13px;">Base de Conocimiento</span>
                <span class="badge success">{len(knowledge_docs)} Docs</span>
              </div>
            </div>
          </div>
        </div>
    </div>
  </div>
"""

    # 1b. TAB PANEL: CONVERSACIONES
    cw_base_url = chatwoot_config.get("base_url") or "https://app.chatwoot.com"
    if chatwoot_enabled:
        cw_url = cw_base_url.rstrip("/")
        cw_account = chatwoot_config.get("account_id", "")
        # Render Chatwoot Iframe
        conversations_panel_html = f"""
        <div id="panel-conversations" class="tab-panel">
          <div class="card" style="margin-top: 20px; text-align: center; padding: 60px 20px;">
            <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" style="width: 48px; height: 48px; color: var(--primary); margin-bottom: 16px;"><path d="M21 15a4 4 0 0 1-4 4H8l-5 3V7a4 4 0 0 1 4-4h10a4 4 0 0 1 4 4v8Z"/></svg>
            <h2>Bandeja de Chatwoot Activa</h2>
            <p style="color: var(--muted); margin-bottom: 24px; max-width: 500px; margin-left: auto; margin-right: auto;">Has configurado Chatwoot como tu bandeja principal. Todos los mensajes se están enviando directamente a tu cuenta externa. Haz clic en el botón de abajo para responder a tus clientes.</p>
            <a href="{html.escape(cw_url)}/app/accounts/{html.escape(cw_account)}/dashboard" target="_blank" class="btn" style="text-decoration: none; display: inline-flex; font-size: 15px; padding: 12px 24px;">
              Ir a Chatwoot <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" style="width:18px; height:18px; margin-left:8px;"><path d="M18 13v6a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2V8a2 2 0 0 1 2-2h6"/><polyline points="15 3 21 3 21 9"/><line x1="10" y1="14" x2="21" y2="3"/></svg>
            </a>
          </div>
        </div>
        """
    else:
        conversations_panel_html = f"""
        <div id="panel-conversations" class="tab-panel">
          <div class="chatwoot-layout">
            <div class="chat-sidebar">
              <div class="chat-sidebar-header">
                Chats activos
                <span class="badge">{len(threads)}</span>
              </div>
              <div class="chat-list">
                {thread_items}
              </div>
            </div>
            <div class="chat-main">
              {chat_header}
              <div class="chat-messages">
                {bubble_html}
              </div>
            </div>
            {crm_sidebar}
          </div>
        </div>
        """

    body_html += conversations_panel_html + f"""
    <!-- 1c. TAB PANEL: CRM -->
    <div id="panel-crm" class="tab-panel">
      <div class="card" style="margin-bottom:20px;">
        <div class="card-header">
          <h2>CRM de Leads</h2>
          <p>Mueve prospectos de no cualificados a calificados conforme avanza la venta.</p>
        </div>
      </div>
      <div class="tabs" style="margin-bottom:16px;">
        {crm_tabs_html}
      </div>
      <section class="table-wrap" style="background:white; border:1px solid var(--line); border-radius:12px; overflow:hidden; box-shadow:var(--shadow);">
        <table style="width:100%; border-collapse:collapse; text-align:left;">
          <thead>
            <tr style="background:#f8fafc; border-bottom:1px solid var(--line);">
              <th style="padding:14px 16px; font-weight:600; color:var(--muted); font-size:13px;">Prospecto</th>
              <th style="padding:14px 16px; font-weight:600; color:var(--muted); font-size:13px;">Negocio</th>
              <th style="padding:14px 16px; font-weight:600; color:var(--muted); font-size:13px;">Estado</th>
              <th style="padding:14px 16px; font-weight:600; color:var(--muted); font-size:13px;">Actualizado</th>
              <th style="padding:14px 16px; font-weight:600; color:var(--muted); font-size:13px;">Acciones</th>
            </tr>
          </thead>
          <tbody>
            {crm_rows}
          </tbody>
        </table>
      </section>
    </div>
    
    <!-- 2. TAB PANEL: WHATSAPP -->
    <div id="panel-whatsapp" class="tab-panel">
      <div class="grid-cards">
        <div class="panel">
          <div class="card">
            <div class="card-header">
              <h2>Meta Embedded Signup (Recomendado)</h2>
              <p>Vincula el número de WhatsApp de tu empresa de forma guiada y segura a través de Asistto como Tech Provider.</p>
            </div>
            
            <div style="margin:20px 0; padding:16px; background:#f0fdfa; border:1px solid #99f6e4; border-radius:8px;">
              <span class="bold-text" style="color:var(--primary-dark); font-size:14px; display:block; margin-bottom:6px;">Instrucciones para Embedded Signup:</span>
              <ol style="margin:0; padding-left:20px; font-size:13px; color:#1e293b; line-height:1.6;">
                <li>Haz clic en el botón verde "Abrir Embedded Signup" abajo.</li>
                <li>Inicia sesión con tu cuenta de negocio de Facebook.</li>
                <li>Selecciona el portafolio comercial de tu negocio y el WABA o número a vincular.</li>
                <li>Una vez completado, los datos de IDs se llenarán automáticamente en el formulario de la derecha. Revísalos y presiona "Guardar conexión".</li>
              </ol>
            </div>
            
            <div style="margin-top:20px;">
              <button class="btn whatsapp-btn" type="button" id="launchMetaSignup">Abrir Embedded Signup de Meta</button>
            </div>
            <div id="metaSignupStatus" class="sync-status">Esperando inicio de vinculación...</div>
          </div>
        </div>
        
        <div>
          <div class="card">
            <div class="card-header">
              <h2>Detalles de la Conexión</h2>
              <p>Completa o edita las credenciales de conexión del bot.</p>
            </div>
            <form method="post" action="/client/bots/{bot_id}/whatsapp/connect">
              <label>Authorization Code de Meta (si aplica)</label>
              <input id="metaAuthCode" name="authorization_code" autocomplete="off" placeholder="Llenado automáticamente">
              
              <label>Access Token (Manual o Temporal)</label>
              <div class="password-wrapper">
                <input type="password" name="access_token" placeholder="EAW..." autocomplete="new-password" value="{html.escape(access_token_val)}">
                <button type="button" class="password-toggle" onclick="togglePasswordVisibility(this)">
                  <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round" style="width: 18px; height: 18px;"><path d="M1 12s4-8 11-8 11 8 11 8-4 8-11 8-11-8-11-8z"></path><circle cx="12" cy="12" r="3"></circle></svg>
                </button>
              </div>
              
              <label>Business Portfolio ID</label>
              <input id="metaBusinessId" name="business_id" value="{html.escape(wa_info.get("business_id") or "")}">
              
              <label>WABA ID (WhatsApp Business Account)</label>
              <input id="metaWabaId" name="waba_id" value="{html.escape(wa_info.get("waba_id") or "")}">
              
              <label>Phone Number ID</label>
              <input id="metaPhoneId" name="phone_number_id" value="{html.escape(wa_info.get("phone_number_id") or "")}" required>
              
              <label>Número de WhatsApp Visible</label>
              <input id="metaDisplayPhone" name="display_phone_number" value="{html.escape(wa_info.get("display_phone_number") or "")}" placeholder="+52...">
              
              <div style="margin-top:20px;">
                <button class="btn primary-btn" type="submit" {"disabled" if session["role"] == "client_viewer" else ""}>Guardar conexión cifrada</button>
              </div>
            </form>
          </div>
        </div>
      </div>
      
      <script async defer crossorigin="anonymous" src="https://connect.facebook.net/en_US/sdk.js"></script>
      <script>
        (() => {{
          const app_id = "{config.META_APP_ID or ""}";
          const config_id = "{config.META_CONFIG_ID or ""}";
          const status = document.getElementById("metaSignupStatus");
          
          window.fbAsyncInit = function() {{
            if (!app_id) return;
            FB.init({{ appId: app_id, cookie: true, xfbml: true, version: 'v19.0' }});
          }};
          
          window.addEventListener("message", (event) => {{
            if (!event.origin.endsWith("facebook.com")) return;
            let data = event.data;
            try {{ if (typeof data === "string") data = JSON.parse(data); }} catch (_) {{ return; }}
            const payload = data?.data || data;
            if (payload?.phone_number_id) document.getElementById("metaPhoneId").value = payload.phone_number_id;
            if (payload?.waba_id) document.getElementById("metaWabaId").value = payload.waba_id;
            if (payload?.business_id) document.getElementById("metaBusinessId").value = payload.business_id;
            if (payload?.display_phone_number) document.getElementById("metaDisplayPhone").value = payload.display_phone_number;
            if (payload?.phone_number_id) status.innerHTML = "<span class='sync-status ok'>Datos recibidos de Meta. Revisa y pulsa 'Guardar conexión'.</span>";
          }});
          
          document.getElementById("launchMetaSignup")?.addEventListener("click", () => {{
            if (!window.FB) {{
              status.innerHTML = "<span class='sync-status err'>Meta SDK no está listo todavía. Inténtalo en un momento.</span>";
              return;
            }}
            FB.login((response) => {{
              const code = response?.authResponse?.code;
              if (code) {{
                document.getElementById("metaAuthCode").value = code;
                status.innerHTML = "<span class='sync-status ok'>Vínculo inicial exitoso. Completa y guarda.</span>";
              }} else {{
                status.innerHTML = "<span class='sync-status err'>Meta canceló el flujo o no regresó código.</span>";
              }}
            }}, {{
              config_id: config_id,
              response_type: "code",
              override_default_response_type: true,
              extras: {{ 
                featureType: "whatsapp_business_app_onboarding",
                sessionInfoVersion: "3",
                setup: {{}}
              }}
            }});
          }});
        }})();
      </script>
    </div>
    
    <!-- 3. TAB PANEL: COMPORTAMIENTO (PROMPT) -->
    <div id="panel-prompt" class="tab-panel">
      <form id="behaviorForm" method="post" action="/client/bots/{bot_id}/prompt/save">
        <div class="prompt-workspace">
        <div class="card">
          <div class="card-header">
            <h2>Agente PBD con IA</h2>
            <p>Escribe qué quieres que haga tu bot en WhatsApp y la IA generará el prompt final integrando tu base de conocimientos.</p>
          </div>
          <div style="margin-bottom:14px;">
            <label>Describe el comportamiento que buscas</label>
            <textarea id="aiPromptInstruction" style="min-height: 120px;" placeholder="Ej. Eres un bot de la Clínica Dental Smile. Tu objetivo es agendar citas de diagnóstico dental. Saluda amablemente, solicita nombre completo y busca un espacio libre. Si tienen urgencias de dolor fuerte, pásalo a humano de inmediato..."></textarea>
          </div>
          <div>
            <button class="btn primary-btn" type="button" id="btnAssistPrompt" onclick="requestAIPrompt()">
              Diseñar comportamiento con PBD
            </button>
            <span id="aiAssistLoader" style="display:none; margin-left:12px; font-size:13px; color:var(--primary); font-weight:600;">Generando prompt óptimo...</span>
          </div>
          
          <div id="promptPreviewBlock" style="margin-top:24px; display:none;">
            <div style="display:flex; justify-content:space-between; align-items:center; margin-bottom:8px;">
              <span class="bold-text" style="font-size:13px; color:var(--primary-dark);">Vista previa del prompt generado</span>
              <button class="btn secondary" style="padding:4px 10px; font-size:11px;" onclick="applyGeneratedPrompt()">Aplicar Preview</button>
            </div>
            
            <label>01 - Constitución (Preview)</label>
            <textarea id="aiConstitutionPreview" style="min-height: 120px; background:#f8fafc; font-family:monospace; font-size:12px; border-color:var(--primary-light); margin-bottom: 12px;" readonly></textarea>
            
            <label>02 - Especificaciones (Preview)</label>
            <textarea id="aiSpecsPreview" style="min-height: 120px; background:#f8fafc; font-family:monospace; font-size:12px; border-color:var(--primary-light); margin-bottom: 12px;" readonly></textarea>
            
            <label>03 - Suite de Pruebas (Preview)</label>
            <textarea id="aiTestSuitePreview" style="min-height: 120px; background:#f8fafc; font-family:monospace; font-size:12px; border-color:var(--primary-light); margin-bottom: 12px;" readonly></textarea>

            <label>04 - Master Prompt (Preview)</label>
            <textarea id="aiPromptPreview" style="min-height: 240px; background:#f8fafc; font-family:monospace; font-size:12px; border-color:var(--primary-light);" readonly></textarea>
          </div>
            <div style="margin-top:24px; border-top: 1px solid var(--border-color); padding-top: 24px;">
              <h3 style="margin-bottom: 12px; font-size: 14px;">Documentos de Referencia PBD</h3>
              
              <label>01 - Constitución (Verdad Absoluta)</label>
              <textarea id="activeConstitutionEditor" name="pbd_constitution" style="min-height: 120px; width: 100%; font-family:monospace; font-size:12px; line-height:1.5; padding: 12px; border: 1px solid var(--border-color); border-radius: 4px; box-sizing: border-box; margin-bottom: 16px;">{html.escape(current_constitution)}</textarea>
  
              <label>02 - Especificaciones (Flujos y Datos)</label>
              <textarea id="activeSpecsEditor" name="pbd_specs" style="min-height: 120px; width: 100%; font-family:monospace; font-size:12px; line-height:1.5; padding: 12px; border: 1px solid var(--border-color); border-radius: 4px; box-sizing: border-box; margin-bottom: 16px;">{html.escape(current_specs)}</textarea>
  
              <label>03 - Suite de Pruebas (Casos de uso)</label>
              <textarea id="activeTestSuiteEditor" name="pbd_test_suite" style="min-height: 120px; width: 100%; font-family:monospace; font-size:12px; line-height:1.5; padding: 12px; border: 1px solid var(--border-color); border-radius: 4px; box-sizing: border-box; margin-bottom: 16px;">{html.escape(current_test_suite)}</textarea>

              <div style="margin-top:4px;">
                <button class="btn secondary" type="submit" {"disabled" if session["role"] == "client_viewer" else ""}>Guardar documentos PBD</button>
              </div>
            </div>
          </div>
          
          <div class="card">
            <div class="card-header">
              <h2>Estructura del Comportamiento (PBD)</h2>
              <p>Puedes editar el comportamiento manualmente o usar el Asistente de IA a la izquierda.</p>
            </div>
            
            <label>04 - Master Prompt (Código del Bot)</label>
            <textarea id="activePromptEditor" name="prompt" style="min-height: 520px; width: 100%; font-family:monospace; font-size:12.5px; line-height:1.5; padding: 12px; border: 1px solid var(--border-color); border-radius: 4px; box-sizing: border-box;" placeholder="System prompt...">{html.escape(current_prompt_content)}</textarea>
            
            <div style="margin-top:16px;">
              <button class="btn" type="submit" {"disabled" if session["role"] == "client_viewer" else ""}>Publicar comportamiento</button>
            </div>
          </div>
        </div>
      </form>
      
      <script>
        document.getElementById("behaviorForm")?.addEventListener("submit", function() {{
          if (window.promptEditor?.codemirror) {{
            window.promptEditor.codemirror.save();
          }}
        }});

        async function requestAIPrompt() {{
          const instruction = document.getElementById("aiPromptInstruction").value.trim();
          const current = window.promptEditor ? window.promptEditor.value() : document.getElementById("activePromptEditor").value;
          if (!instruction) {{
            alert("Por favor escribe una descripción para el comportamiento del bot.");
            return;
          }}
          
          const loader = document.getElementById("aiAssistLoader");
          const btn = document.getElementById("btnAssistPrompt");
          const previewBlock = document.getElementById("promptPreviewBlock");
          const previewArea = document.getElementById("aiPromptPreview");
          const constPreview = document.getElementById("aiConstitutionPreview");
          const specsPreview = document.getElementById("aiSpecsPreview");
          const testPreview = document.getElementById("aiTestSuitePreview");
          
          loader.style.display = "inline";
          btn.disabled = true;
          
          const formData = new FormData();
          formData.append("instruction", instruction);
          formData.append("current_prompt", current);
          formData.append("pbd_constitution", document.getElementById("activeConstitutionEditor").value);
          formData.append("pbd_specs", document.getElementById("activeSpecsEditor").value);
          formData.append("pbd_test_suite", document.getElementById("activeTestSuiteEditor").value);
          
          try {{
            const response = await fetch("/client/bots/{bot_id}/prompt/assist", {{
              method: "POST",
              body: formData
            }});
            if (!response.ok) {{
              const err = await response.json();
              throw new Error(err.detail || "Error generando el prompt");
            }}
            const data = await response.json();
            if (data.ok) {{
              previewArea.value = data.prompt;
              constPreview.value = data.pbd_constitution || "";
              specsPreview.value = data.pbd_specs || "";
              testPreview.value = data.pbd_test_suite || "";
              previewBlock.style.display = "block";
            }} else {{
              alert("Error: " + data.error);
            }}
          }} catch (e) {{
            alert(e.message);
          }} finally {{
            loader.style.display = "none";
            btn.disabled = false;
          }}
        }}
        
        function applyGeneratedPrompt() {{
          const generated = document.getElementById("aiPromptPreview").value;
          const generatedConst = document.getElementById("aiConstitutionPreview").value;
          const generatedSpecs = document.getElementById("aiSpecsPreview").value;
          const generatedTest = document.getElementById("aiTestSuitePreview").value;

          if (window.promptEditor) {{
            window.promptEditor.value(generated);
          }} else {{
            document.getElementById("activePromptEditor").value = generated;
          }}
          
          document.getElementById("activeConstitutionEditor").value = generatedConst;
          document.getElementById("activeSpecsEditor").value = generatedSpecs;
          document.getElementById("activeTestSuiteEditor").value = generatedTest;

          document.getElementById("promptPreviewBlock").style.display = "none";
          alert("Documentos aplicados a los campos. Haz clic en 'Publicar comportamiento' para guardarlo definitivamente.");
        }}
      </script>
      <style>
        /* Estilos para hacer EasyMDE más presentable y acorde al diseño */
        .editor-toolbar {{
          border: 1px solid var(--border-color) !important;
          border-bottom: none !important;
          border-radius: 8px 8px 0 0 !important;
          opacity: 1 !important;
          background: #f8fafc;
        }}
        .CodeMirror {{
          border: 1px solid var(--border-color) !important;
          border-radius: 0 0 8px 8px !important;
          font-family: 'Inter', system-ui, sans-serif !important;
          font-size: 13.5px !important;
          line-height: 1.6 !important;
          color: var(--text-main) !important;
        }}
        /* Reducir el tamaño gigantesco de los headers dentro del editor */
        .cm-s-easymde .cm-header-1 {{ font-size: 1.35em !important; font-weight: 700 !important; color: var(--primary-dark) !important; }}
        .cm-s-easymde .cm-header-2 {{ font-size: 1.25em !important; font-weight: 700 !important; color: var(--primary-dark) !important; }}
        .cm-s-easymde .cm-header-3 {{ font-size: 1.15em !important; font-weight: 600 !important; }}
        .cm-s-easymde .cm-header-4 {{ font-size: 1.05em !important; font-weight: 600 !important; }}
        .editor-preview {{
          font-family: 'Inter', system-ui, sans-serif !important;
          font-size: 14px !important;
          line-height: 1.6 !important;
          background: #f8fafc !important;
        }}
        .editor-preview h1 {{ font-size: 1.5em !important; margin-bottom: 0.5em; border-bottom: 1px solid var(--border-color); padding-bottom: 8px; }}
        .editor-preview h2 {{ font-size: 1.3em !important; margin-bottom: 0.5em; }}
        .editor-preview h3 {{ font-size: 1.1em !important; margin-bottom: 0.5em; }}
        .editor-toolbar button.active, .editor-toolbar button:hover {{
          background: var(--bg-hover) !important;
          border-radius: 4px;
        }}
      </style>
      <link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/easymde/dist/easymde.min.css">
      <script src="https://cdn.jsdelivr.net/npm/easymde/dist/easymde.min.js"></script>
      <script>
        document.addEventListener("DOMContentLoaded", function() {{
          window.promptEditor = new EasyMDE({{
            element: document.getElementById('activePromptEditor'),
            spellChecker: false,
            minHeight: "350px"
          }});
        }});
      </script>
    </div>
    
    <!-- 4. TAB PANEL: HORARIOS -->
    <div id="panel-hours" class="tab-panel">
      <div class="card" style="max-width:840px;">
        <div class="card-header">
          <h2>Horarios de Atención y Agendado</h2>
          <p>Define las horas y días laborales del bot en las que está permitido programar llamadas o citas en tu calendario. Fuera de estas horas, el bot avisará de forma determinista al cliente.</p>
        </div>
        <form method="post" action="/client/bots/{bot_id}/hours">
          <div class="checkbox-group" style="margin-bottom:20px;">
            <input type="checkbox" name="enabled" id="hoursEnabledToggle" {"checked" if hours_enabled else ""}>
            <label for="hoursEnabledToggle" style="margin:0; font-size:14px; font-weight:600; cursor:pointer;">Activar restricción de horarios de atención para agendado</label>
          </div>
          
          <div style="border: 1px solid var(--line); border-radius:8px; overflow:hidden; background:white;">
            <div class="weekday-row" style="background:#f8fafc; font-weight:700; font-size:11px; text-transform:uppercase; color:var(--muted); border-bottom: 2px solid var(--line);">
              <div>Día</div>
              <div>¿Abierto?</div>
              <div>Hora Inicio</div>
              <div>Hora Cierre</div>
            </div>
            
            {_render_hours_rows(hours_config)}
          </div>
          
          <div style="margin-top:20px;">
            <button class="btn primary-btn" type="submit" {"disabled" if session["role"] == "client_viewer" else ""}>Guardar horarios</button>
          </div>
        </form>
      </div>
    </div>
    
    <!-- 5. TAB PANEL: ESCALADO -->
    <div id="panel-escalate" class="tab-panel">
      <div class="card" style="max-width: 720px;">
        <div class="card-header">
          <h2>Reglas de Escalado Humano</h2>
          <p>Configura las condiciones bajo las cuales el bot debe detenerse y marcar la conversación para que un asesor humano tome el control en WhatsApp.</p>
        </div>
        
        <form method="post" action="/client/bots/{bot_id}/escalation">
          <div class="checkbox-group" style="margin-bottom:18px;">
            <input type="checkbox" name="enabled" id="escalateEnabledToggle" {"checked" if escalate_enabled else ""}>
            <label for="escalateEnabledToggle" style="margin:0; font-size:14px; font-weight:600; cursor:pointer;">Activar escalación humana de conversaciones</label>
          </div>
          
          <div style="margin-bottom:20px;">
            <label>Palabras clave o frases de activación (una por línea o separadas por comas)</label>
            <textarea name="keywords" style="min-height: 120px;" placeholder="Ej. queja, hablar con humano, operador, urgente, reclamo, costo extra">{html.escape(", ".join(escalate_config.get("keywords", [])))}</textarea>
            <p class="muted-text" style="margin-top:4px;">Si el usuario escribe cualquiera de estas frases, el bot registrará la escalación y se marcará en color rojo como 'pendiente' en el panel.</p>
          </div>
          
          <div class="checkbox-group">
            <input type="checkbox" name="escalate_on_media" id="escalateOnMedia" {"checked" if escalate_config.get("escalate_on_media", True) else ""}>
            <label for="escalateOnMedia" style="margin:0; font-weight:500; cursor:pointer;">Escalar automáticamente si el cliente envía un archivo (imagen, audio, pdf, etc.)</label>
          </div>
          
          <div style="margin-top:20px;">
            <button class="btn primary-btn" type="submit" {"disabled" if session["role"] == "client_viewer" else ""}>Guardar reglas de escalado</button>
          </div>
        </form>
      </div>
    </div>
    
    <!-- 6. TAB PANEL: CONOCIMIENTO -->
    <div id="panel-knowledge" class="tab-panel">
      <div class="grid-cards">
        <div class="card">
          <div class="card-header">
            <h2>Base de Conocimiento Activa</h2>
            <p>Información que el bot puede leer como contexto para contestar preguntas sobre precios, servicios, ubicaciones o preguntas frecuentes.</p>
          </div>
          {_render_knowledge_table(knowledge_docs, bot_id, session["role"])}
        </div>
        
        <div>
          <div class="card">
            <div class="card-header">
              <h2>Agregar Documento o Archivo</h2>
              <p>Sube preguntas frecuentes, menús de servicio o políticas operativas. Puedes escribir texto directamente o subir un archivo.</p>
            </div>
            <form method="post" action="/client/bots/{bot_id}/knowledge" enctype="multipart/form-data">
              <label>Título del Documento (Opcional si subes archivo)</label>
              <input name="title" placeholder="Preguntas Frecuentes Dentales">
              
              <label>Contenido del Conocimiento (Escribe texto...)</label>
              <textarea name="content" style="min-height: 140px;" placeholder="Ej. &#10;¿Tienen estacionamiento? Sí, gratuito en plaza.&#10;¿Precios de Limpieza? Desde $500 MXN.&#10;¿Aceptan tarjeta? Sí, Visa y Mastercard."></textarea>
              
              <label style="margin-top:14px; display:block; font-weight:600;">...o Sube un Archivo (PDF, Word, MD, Excel, CSV)</label>
              <div style="display:flex; align-items:center; gap:10px; margin-top:6px;">
                <input type="file" id="knowledgeFileInput" name="file" accept=".pdf,.docx,.xlsx,.csv,.md,.txt" style="font-size:13px; color:var(--muted);">
                <button type="button" class="btn secondary" onclick="document.getElementById('knowledgeFileInput').value='';" style="padding:4px 8px; font-size:11px; margin:0; cursor:pointer;">Quitar archivo</button>
              </div>
              
              <div style="margin-top:18px;">
                <button class="btn primary-btn" type="submit" {"disabled" if session["role"] == "client_viewer" else ""}>Guardar documento</button>
              </div>
            </form>
          </div>
        </div>
      </div>
    </div>
    
    <!-- 6b. TAB PANEL: PLANTILLAS -->
    {templates_panel_html}
    
    <!-- 6c. TAB PANEL: CONTACTOS -->
    {contacts_panel_html}
    
    <!-- 6d. TAB PANEL: CAMPAÑAS -->
    {campaigns_panel_html}
    
    <!-- 7. TAB PANEL: INTEGRACIONES -->
    <div id="panel-integrations" class="tab-panel">
      <div class="grid-2">
        <div class="card">
          <div class="card-header">
            <h2>Chatwoot (Bandeja Multicanal)</h2>
            <p>Conecta tu cuenta de Chatwoot para responder manualmente a tus clientes cuando la IA transfiere el chat.</p>
          </div>
          
          <div style="margin-bottom:14px; display:flex; justify-content:space-between; align-items:center; padding-bottom:8px; border-bottom:1px solid var(--line);">
            <span class="bold-text">Estado de Integración</span>
            <span class="badge {"success" if chatwoot_enabled else "warning"}">
              {"Conectado" if chatwoot_enabled else "Desconectado"}
            </span>
          </div>
          
          <form method="post" action="/client/bots/{bot_id}/integrations/chatwoot">
            <div class="checkbox-group">
              <input type="checkbox" name="enabled" id="chatwootToggle" {"checked" if chatwoot_enabled else ""}>
              <label for="chatwootToggle" style="margin:0; font-weight:600; cursor:pointer;">Habilitar Sincronización con Chatwoot</label>
            </div>
            
            <label>URL Base de Chatwoot</label>
            <input name="base_url" placeholder="https://app.chatwoot.com" value="{html.escape(chatwoot_config.get("base_url") or "")}">
            
            <div style="display:flex; gap:10px;">
              <div style="flex:1;">
                <label>Account ID</label>
                <input name="account_id" placeholder="Ej. 1" value="{html.escape(chatwoot_config.get("account_id") or "")}">
              </div>
              <div style="flex:1;">
                <label>Inbox ID</label>
                <input name="inbox_id" placeholder="Ej. 12" value="{html.escape(chatwoot_config.get("inbox_id") or "")}">
              </div>
            </div>
            
            <label>User API Token</label>
            <div class="password-wrapper">
              <input type="password" name="api_token" placeholder="Copia aquí el token" autocomplete="new-password" value="{html.escape(chatwoot_secrets.get("api_token") or "")}">
              <button type="button" class="password-toggle" onclick="togglePasswordVisibility(this)">
                <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round" style="width: 18px; height: 18px;"><path d="M1 12s4-8 11-8 11 8 11 8-4 8-11 8-11-8-11-8z"></path><circle cx="12" cy="12" r="3"></circle></svg>
              </button>
            </div>
            
            <div style="margin-top:20px; display:flex; gap:10px;">
              <button class="btn primary-btn" type="submit" {"disabled" if session["role"] == "client_viewer" else ""}>Guardar Chatwoot</button>
            </div>
          </form>
        </div>

        <div class="card">
          <div class="card-header">
            <h2>Agenda (Google Calendar)</h2>
            <p>Conecta la cuenta de Google Calendar de tu negocio para agendar de forma autónoma.</p>
          </div>
          
          <div style="margin-bottom:14px; display:flex; justify-content:space-between; align-items:center; padding-bottom:8px; border-bottom:1px solid var(--line);">
            <span class="bold-text">Habilidad de Calendario</span>
            <span class="badge {"success" if calendar_status.get("enabled") else "warning"}">
              {"Activa" if calendar_status.get("enabled") else "Inactiva"}
            </span>
          </div>
          
          <form method="post" action="/client/bots/{bot_id}/integrations/calendar">
            <div class="checkbox-group">
              <input type="checkbox" name="enabled" id="calendarSkillToggle" {"checked" if calendar_status.get("skill_enabled", True) else ""}>
              <label for="calendarSkillToggle" style="margin:0; font-weight:600; cursor:pointer;">Habilitar reserva automática</label>
            </div>
            
            <label>Google Client ID</label>
            <input name="client_id" placeholder="12345-abcde.apps.googleusercontent.com" value="{html.escape(calendar_config.get("client_id") or "")}">
            
            <label>Google Client Secret</label>
            <div class="password-wrapper">
              <input type="password" name="client_secret" placeholder="********" autocomplete="new-password" value="{"" if not calendar_status.get("secret_client_secret_saved") else "********"}">
              <button type="button" class="password-toggle" onclick="togglePasswordVisibility(this)">
                <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round" style="width: 18px; height: 18px;"><path d="M1 12s4-8 11-8 11 8 11 8-4 8-11 8-11-8-11-8z"></path><circle cx="12" cy="12" r="3"></circle></svg>
              </button>
            </div>
            
            <label>Google Refresh Token</label>
            <div class="password-wrapper">
              <input type="password" name="refresh_token" placeholder="1//0..." autocomplete="new-password" value="{"" if not calendar_status.get("secret_refresh_token_saved") else "********"}">
              <button type="button" class="password-toggle" onclick="togglePasswordVisibility(this)">
                <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round" style="width: 18px; height: 18px;"><path d="M1 12s4-8 11-8 11 8 11 8-4 8-11 8-11-8-11-8z"></path><circle cx="12" cy="12" r="3"></circle></svg>
              </button>
            </div>
            
            <label>Google Calendar ID</label>
            <input name="calendar_id" placeholder="ej. primary o email@gmail.com" value="{html.escape(calendar_config.get("calendar_id") or "primary")}">
            
            <label>Zona Horaria</label>
            <input name="timezone" placeholder="America/Chihuahua" value="{html.escape(calendar_config.get("timezone") or config.GOOGLE_CALENDAR_TIMEZONE)}">
            
            <div style="margin-top:20px; display:flex; gap:10px;">
              <button class="btn primary-btn" type="submit" {"disabled" if session["role"] == "client_viewer" else ""}>Guardar Agenda</button>
            </div>
          </form>
        </div>
        
        <div class="card">
          <div class="card-header">
            <h2>Variables de Entorno (API / Webhook)</h2>
            <p>Configura variables dinámicas (Ej. <code>MERCADOPAGO_TOKEN</code>) para usarlas en llamadas externas reemplazando su nombre con llaves: <code>{{MERCADOPAGO_TOKEN}}</code>.</p>
          </div>
          
          <form method="post" action="/client/bots/{bot_id}/integrations/api">
            <label>URL Base del API (Opcional)</label>
            <input name="base_url" placeholder="https://api.miclinicadental.com/v1" value="{html.escape(api_config.get("base_url") or "")}">
            
            <div style="margin-top: 16px;">
              <label>Variables de Entorno</label>
              <div id="envVarsContainer" style="display:flex; flex-direction:column; gap:8px;">
                {env_rows_html}
              </div>
              <button type="button" class="btn secondary" style="margin-top:12px; font-size:12px;" onclick="addEnvVarRow()">+ Agregar Variable</button>
            </div>
            
            <div style="margin-top:24px;">
              <button class="btn primary-btn" type="submit" {"disabled" if session["role"] == "client_viewer" else ""}>Guardar Variables</button>
            </div>
          </form>
          
          <script>
            function addEnvVarRow() {{
              const container = document.getElementById('envVarsContainer');
              const row = document.createElement('div');
              row.className = 'env-var-row';
              row.style.cssText = 'display:flex; gap:8px; align-items:center; margin-top:4px;';
              row.innerHTML = `
                <input name="env_key" placeholder="KEY (Ej. API_KEY)" style="flex:1; margin:0;" required>
                <div class="password-wrapper" style="flex:2;">
                  <input type="password" name="env_val" placeholder="Valor del secreto" autocomplete="new-password" style="margin:0;" required>
                  <button type="button" class="password-toggle" onclick="togglePasswordVisibility(this)">
                    <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round" style="width: 18px; height: 18px;"><path d="M1 12s4-8 11-8 11 8 11 8-4 8-11 8-11-8-11-8z"></path><circle cx="12" cy="12" r="3"></circle></svg>
                  </button>
                </div>
                <button type="button" class="btn secondary" style="padding:8px 12px; color:var(--red);" onclick="this.parentElement.remove()">X</button>
              `;
              container.appendChild(row);
            }}
          </script>
        </div>

        <div class="card">
          <div class="card-header">
            <h2>Enrutamiento Externo y Reglas de Omisión</h2>
            <p>Define números de teléfono de tu equipo (vendedores/agentes) para que el Bot de IA los ignore y desvíe sus mensajes automáticamente a tu Webhook.</p>
          </div>
          
          <div style="margin-bottom:14px; display:flex; justify-content:space-between; align-items:center; padding-bottom:8px; border-bottom:1px solid var(--line);">
            <span class="bold-text">Estado de Enrutamiento</span>
            <span class="badge {"success" if routing_enabled else "warning"}">
              {"Activo" if routing_enabled else "Inactivo"}
            </span>
          </div>
          
          <form method="post" action="/client/bots/{bot_id}/integrations/routing">
            <div class="checkbox-group">
              <input type="checkbox" name="enabled" id="routingEnabledToggle" {"checked" if routing_enabled else ""}>
              <label for="routingEnabledToggle" style="margin:0; font-weight:600; cursor:pointer;">Habilitar enrutamiento externo</label>
            </div>
            
            <label>URL del Webhook de Destino</label>
            <input name="webhook_url" placeholder="https://mi-sistema.com/webhook-whatsapp" value="{html.escape(routing_webhook_url or "")}">
            
            <label>Números de Teléfono a Excluir (Separados por comas)</label>
            <textarea name="phone_numbers" placeholder="Ej. 5216861234567, 5215559876543" style="min-height:80px; font-family:monospace; font-size:12.5px;">{html.escape(routing_phone_numbers or "")}</textarea>
            <span class="muted-text" style="font-size:11px; display:block; margin-top:-6px; margin-bottom:12px;">Los mensajes entrantes de estos números serán desviados al Webhook y omitidos por la IA.</span>
            
            <div class="checkbox-group" style="margin-bottom:16px;">
              <input type="checkbox" name="save_history" id="routingSaveHistoryToggle" {"checked" if routing_save_history else ""}>
              <label for="routingSaveHistoryToggle" style="margin:0; font-weight:500; cursor:pointer; font-size:13px;">Registrar mensajes desviados en el historial del chat (lectura)</label>
            </div>
            
            <label>Token de Autenticación del Webhook (Opcional)</label>
            <div class="password-wrapper">
              <input type="password" name="webhook_auth_token" placeholder="********" autocomplete="new-password" value="{"" if not routing_token_saved else "********"}">
              <button type="button" class="password-toggle" onclick="togglePasswordVisibility(this)">
                <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round" style="width: 18px; height: 18px;"><path d="M1 12s4-8 11-8 11 8 11 8-4 8-11 8-11-8-11-8z"></path><circle cx="12" cy="12" r="3"></circle></svg>
              </button>
            </div>
            <span class="muted-text" style="font-size:11px; display:block; margin-top:-6px; margin-bottom:12px;">Se enviará en la cabecera <code>X-Asistto-Secret-Token</code> para validar la llamada.</span>
            
            <div style="margin-top:20px; display:flex; gap:10px;">
              <button class="btn primary-btn" type="submit" {"disabled" if session["role"] == "client_viewer" else ""}>Guardar Enrutamiento</button>
            </div>
          </form>
        </div>
      </div>
      
      <!-- Skill Templates Catalog -->
      <div style="margin-top: 40px; margin-bottom: 20px; border-top: 1px solid var(--line); padding-top: 30px; width: 100%;">
        <h2 style="font-family:Outfit; font-size:22px; font-weight:700; color:var(--ink);">Catálogo de Habilidades Pre-construidas</h2>
        <p style="color:var(--muted); font-size:13.5px; margin-top:4px;">Instala plantillas listas para usar que automatizan integraciones comunes en este bot.</p>
      </div>
      <div class="grid-3" style="gap: 20px; width: 100%;">
        {_render_skill_templates(bot_id, session)}
      </div>
    </div>
    """
    
    return HTMLResponse(_layout(
        title="Panel",
        body=body_html,
        session=session,
        active_tab=tab,
        notice=notice_html,
        bots_list=bots,
        selected_bot_id=bot_id,
        chatwoot_enabled=chatwoot_enabled,
        cw_base_url=cw_base_url,
        cw_account=cw_account,
    ))

def _render_conversations_table(threads: list) -> str:
    if not threads:
        return '<div style="padding:28px 0; text-align:center; color:var(--muted)">Sin conversaciones registradas todavía.</div>'
        
    rows = ""
    for thread in threads:
        wa_id = thread["wa_id"]
        last_role = thread["last_role"]
        last_content = thread["last_content"] or ""
        role_badge = '<span class="badge" style="padding:2px 6px; font-size:10px;">Cliente</span>' if last_role == "user" else '<span class="badge success" style="padding:2px 6px; font-size:10px;">Bot</span>'
        
        display_name = thread["nombre"] or wa_id
        lead_badge = ""
        status = thread["qualification_status"] or "en_progreso"
        if status == "calificado":
            lead_badge = '<span class="badge success" style="padding:2px 6px; font-size:10px;">Calificado</span>'
        elif status == "descalificado":
            lead_badge = '<span class="badge danger" style="padding:2px 6px; font-size:10px;">Descalificado</span>'
        else:
            lead_badge = '<span class="badge warning" style="padding:2px 6px; font-size:10px;">En progreso</span>'
            
        rows += f"""
        <tr>
          <td style="font-weight:600; font-size:13px; width:180px;">{html.escape(display_name)}<br><span class="muted-text" style="font-size:11px;">{html.escape(wa_id)}</span></td>
          <td style="font-size:13px; max-width:320px; overflow:hidden; text-overflow:ellipsis; white-space:nowrap;">{role_badge} {html.escape(last_content[:100])}</td>
          <td style="width:120px; text-align:right;">{lead_badge}</td>
        </tr>
        """
        
    return f"""
    <div style="overflow-x:auto;">
      <table style="width:100%;">
        <thead>
          <tr>
            <th>Contacto</th>
            <th>Último mensaje</th>
            <th style="text-align:right;">CRM</th>
          </tr>
        </thead>
        <tbody>
          {rows}
        </tbody>
      </table>
    </div>
    """

def _render_hours_rows(config_data: dict) -> str:
    days = [
        ("lunes", "Lunes"),
        ("martes", "Martes"),
        ("miercoles", "Miércoles"),
        ("jueves", "Jueves"),
        ("viernes", "Viernes"),
        ("sabado", "Sábado"),
        ("domingo", "Domingo"),
    ]
    
    rows_html = ""
    for key, label in days:
        day_cfg = config_data.get(key, {})
        is_open = day_cfg.get("open", key not in ("sabado", "domingo"))
        start_time = day_cfg.get("start", "09:00")
        end_time = day_cfg.get("end", "18:00")
        
        rows_html += f"""
        <div class="weekday-row">
          <div class="bold-text" style="font-size:14px;">{label}</div>
          <div>
            <input type="checkbox" name="open_{key}" {"checked" if is_open else ""} style="width:18px; height:18px; accent-color:var(--primary); cursor:pointer;">
          </div>
          <div>
            <input type="text" name="start_{key}" value="{html.escape(start_time)}" style="max-width:110px; text-align:center; padding:6px 10px;" placeholder="09:00">
          </div>
          <div>
            <input type="text" name="end_{key}" value="{html.escape(end_time)}" style="max-width:110px; text-align:center; padding:6px 10px;" placeholder="18:00">
          </div>
        </div>
        """
    return rows_html

def _render_knowledge_table(docs: list, bot_id: int, role: str) -> str:
    if not docs:
        return '<div style="padding:28px 0; text-align:center; color:var(--muted)">No tienes documentos de conocimiento activos. Agrega uno a la derecha.</div>'
        
    rows = ""
    for doc in docs:
        title = doc["title"]
        content = doc["content"]
        created_at = doc.get("created_at")
        dt_str = created_at.strftime("%d/%m/%Y") if created_at else "-"
        
        archive_form = ""
        if role in ("agency_admin", "client_admin"):
            archive_form = f"""
            <form method="post" action="/client/bots/{bot_id}/knowledge/{doc["id"]}/archive" style="display:inline;">
              <button class="btn secondary" style="padding:5px 10px; font-size:11px; color:var(--red); border-color:rgba(225, 29, 72, 0.2); background:#fff5f5;" type="submit">Archivar</button>
            </form>
            """
            
        rows += f"""
        <tr>
          <td style="font-weight:600; font-size:13.5px; width:220px;">{html.escape(title)}<br><span class="muted-text" style="font-size:11px;">Creado: {dt_str}</span></td>
          <td style="font-size:12.5px; color:var(--muted); max-width:380px; overflow:hidden; text-overflow:ellipsis; white-space:nowrap;">{html.escape(content[:160])}</td>
          <td style="width:100px; text-align:right;">{archive_form}</td>
        </tr>
        """
        
    return f"""
    <div style="overflow-x:auto;">
      <table style="width:100%;">
        <thead>
          <tr>
            <th>Documento</th>
            <th>Previsualización</th>
            <th></th>
          </tr>
        </thead>
        <tbody>
          {rows}
        </tbody>
      </table>
    </div>
    """

# --- POST ENDPOINTS FOR SAVING CLIENT CONFIGURATIONS ---

@router.post("/bots/{bot_id}/whatsapp/connect")
async def client_whatsapp_connect(
    request: Request,
    bot_id: int,
    authorization_code: str = Form(""),
    access_token: str = Form(""),
    business_id: str = Form(""),
    waba_id: str = Form(""),
    phone_number_id: str = Form(...),
    display_phone_number: str = Form(""),
):
    session = _require_client_login(request)
    await _require_bot_editor(session, bot_id)
    try:
        await meta_provider.connect_bot_from_embedded_signup(
            meta_provider.MetaConnectionInput(
                bot_id=bot_id,
                phone_number_id=phone_number_id.strip(),
                display_phone_number=display_phone_number.strip(),
                waba_id=waba_id.strip(),
                business_id=business_id.strip(),
                authorization_code=authorization_code.strip(),
                access_token=access_token.strip(),
            )
        )
        return RedirectResponse(f"/client/app?bot_id={bot_id}&tab=whatsapp&saved=1", status_code=302)
    except Exception as exc:
        log.exception("Error vinculando WhatsApp desde panel cliente")
        return RedirectResponse(f"/client/app?bot_id={bot_id}&tab=whatsapp&saved=err", status_code=302)

@router.post("/bots/{bot_id}/prompt/save")
async def client_prompt_save(
    request: Request,
    bot_id: int,
    prompt: str = Form(...),
    pbd_constitution: str | None = Form(None),
    pbd_specs: str | None = Form(None),
    pbd_test_suite: str | None = Form(None),
):
    session = _require_client_login(request)
    await _require_bot_editor(session, bot_id)
    clean = prompt.strip()
    if not clean:
        raise HTTPException(status_code=400, detail="El prompt no puede estar vacío")
    await db.publish_bot_prompt(bot_id, clean, pbd_constitution, pbd_specs, pbd_test_suite)
    return RedirectResponse(f"/client/app?bot_id={bot_id}&tab=prompt&saved=1", status_code=302)

@router.post("/bots/{bot_id}/prompt/assist", response_class=JSONResponse)
async def client_prompt_assist(
    request: Request,
    bot_id: int,
    instruction: str = Form(...),
    current_prompt: str = Form(""),
    pbd_constitution: str = Form(""),
    pbd_specs: str = Form(""),
    pbd_test_suite: str = Form(""),
):
    session = _require_client_login(request)
    bot = await _require_bot_editor(session, bot_id)
    try:
        knowledge_docs = await db.list_bot_knowledge(bot_id, active_only=True)
        result = await prompt_assistant.assist_prompt(
            bot=bot,
            current_prompt=current_prompt,
            pbd_constitution=pbd_constitution,
            pbd_specs=pbd_specs,
            pbd_test_suite=pbd_test_suite,
            instruction=instruction,
            knowledge_docs=knowledge_docs,
        )
        return JSONResponse(result)
    except prompt_assistant.PromptAssistantError as exc:
        return JSONResponse({"ok": False, "error": str(exc)}, status_code=400)
    except Exception as exc:
        return JSONResponse({"ok": False, "error": prompt_assistant.safe_error(exc)}, status_code=502)

@router.post("/bots/{bot_id}/hours")
async def client_hours_save(request: Request, bot_id: int, enabled: str | None = Form(None)):
    session = _require_client_login(request)
    await _require_bot_editor(session, bot_id)
    
    # Parse form parameters for each weekday
    days = ["lunes", "martes", "miercoles", "jueves", "viernes", "sabado", "domingo"]
    hours_config = {}
    form_data = await request.form()
    
    for day in days:
        is_open = form_data.get(f"open_{day}") == "on"
        start_time = str(form_data.get(f"start_{day}") or "09:00").strip()
        end_time = str(form_data.get(f"end_{day}") or "18:00").strip()
        
        # Validar formato HH:MM
        if not re.match(r"^\d{2}:\d{2}$", start_time) or not re.match(r"^\d{2}:\d{2}$", end_time):
            return RedirectResponse(f"/client/app?bot_id={bot_id}&tab=hours&saved=err", status_code=302)
            
        hours_config[day] = {
            "open": is_open,
            "start": start_time,
            "end": end_time
        }
        
    await db.upsert_bot_skill(
        bot_id=bot_id,
        skill_type="business_hours",
        enabled=enabled == "on",
        config_data=hours_config
    )
    return RedirectResponse(f"/client/app?bot_id={bot_id}&tab=hours&saved=1", status_code=302)

@router.post("/bots/{bot_id}/escalation")
async def client_escalation_save(
    request: Request,
    bot_id: int,
    enabled: str | None = Form(None),
    keywords: str = Form(""),
    escalate_on_media: str | None = Form(None),
):
    session = _require_client_login(request)
    await _require_bot_editor(session, bot_id)
    
    # Process comma-separated or newline-separated keywords
    words_list = []
    for chunk in re.split(r"[,\n]", keywords):
        clean = chunk.strip().lower()
        if clean:
            words_list.append(clean)
            
    escalation_config = {
        "keywords": words_list,
        "escalate_on_media": escalate_on_media == "on"
    }
    
    await db.upsert_bot_skill(
        bot_id=bot_id,
        skill_type="escalation",
        enabled=enabled == "on",
        config_data=escalation_config
    )
    return RedirectResponse(f"/client/app?bot_id={bot_id}&tab=escalate&saved=1", status_code=302)

@router.post("/bots/{bot_id}/knowledge")
async def client_knowledge_create(
    request: Request,
    bot_id: int,
    title: str = Form(""),
    content: str = Form(""),
    file: UploadFile | None = File(None),
):
    session = _require_client_login(request)
    await _require_bot_editor(session, bot_id)
    clean_title = title.strip()
    clean_content = content.strip()
    
    if file and file.filename:
        filename = file.filename
        _, ext = os.path.splitext(filename.lower())
        if ext not in (".pdf", ".docx", ".xlsx", ".csv", ".md", ".txt"):
            return RedirectResponse(f"/client/app?bot_id={bot_id}&tab=knowledge&saved=err_ext", status_code=302)
        try:
            await file.seek(0)
            file_bytes = await file.read()
            if len(file_bytes) > config.KNOWLEDGE_UPLOAD_MAX_BYTES:
                return RedirectResponse(f"/client/app?bot_id={bot_id}&tab=knowledge&saved=err_file_too_large", status_code=302)
            if len(file_bytes) > 0:
                parsed_text = file_parser.parse_file(file_bytes, filename)
                if not parsed_text.strip():
                    raise ValueError("El archivo está vacío o no contiene texto legible.")
                clean_content = parsed_text.strip()
                if not clean_title:
                    clean_title = filename
        except Exception as e:
            log.exception(f"Error parseando archivo: {filename}")
            return RedirectResponse(f"/client/app?bot_id={bot_id}&tab=knowledge&saved=err_parse", status_code=302)
            
    if not clean_title or not clean_content:
        raise HTTPException(status_code=400, detail="El título y el contenido son obligatorios (o sube un archivo válido).")
        
    try:
        await db.create_bot_knowledge(
            bot_id=bot_id,
            title=clean_title,
            content=clean_content
        )
    except Exception as e:
        log.exception(f"Error guardando documento de conocimiento en base de datos: {clean_title}")
        return RedirectResponse(f"/client/app?bot_id={bot_id}&tab=knowledge&saved=err", status_code=302)
        
    return RedirectResponse(f"/client/app?bot_id={bot_id}&tab=knowledge&saved=1", status_code=302)

@router.post("/bots/{bot_id}/knowledge/{knowledge_id}/archive")
async def client_knowledge_archive(request: Request, bot_id: int, knowledge_id: int):
    session = _require_client_login(request)
    await _require_bot_editor(session, bot_id)
    
    success = await db.archive_bot_knowledge(bot_id, knowledge_id)
    if not success:
         raise HTTPException(status_code=404, detail="Documento no encontrado o no pertenece a tu bot.")
         
    return RedirectResponse(f"/client/app?bot_id={bot_id}&tab=knowledge&saved=1", status_code=302)

@router.post("/bots/{bot_id}/toggle-status")
async def client_bot_toggle_status(request: Request, bot_id: int):
    session = _require_client_login(request)
    bot = await _require_bot_editor(session, bot_id)
    
    current_status = bot.get("status") or "active"
    new_status = "paused" if current_status == "active" else "active"
    
    await db.update_bot_status(bot_id, new_status)
    
    referer = request.headers.get("referer") or f"/client/app?bot_id={bot_id}"
    return RedirectResponse(referer, status_code=302)

@router.post("/bots/{bot_id}/integrations/calendar")
async def client_calendar_save(
    request: Request,
    bot_id: int,
    enabled: str | None = Form(None),
    client_id: str = Form(...),
    client_secret: str = Form(""),
    refresh_token: str = Form(""),
    calendar_id: str = Form("primary"),
    timezone: str = Form("America/Chihuahua"),
):
    session = _require_client_login(request)
    await _require_bot_editor(session, bot_id)
    
    # Form data validation
    clean_cid = client_id.strip()
    clean_cal = calendar_id.strip() or "primary"
    clean_tz = timezone.strip() or "America/Chihuahua"
    
    # Search or create bot integration record
    integration = await db.get_active_bot_integration(bot_id, "google_calendar")
    
    config_data = {
        "client_id": clean_cid,
        "calendar_id": clean_cal,
        "timezone": clean_tz
    }
    
    if integration:
        integration_id = int(integration["id"])
        await db.update_bot_integration(
            bot_id=bot_id,
            integration_id=integration_id,
            integration_type="google_calendar",
            name=integration["name"],
            config_data=config_data,
            enabled=enabled == "on"
        )
    else:
        integration_id = await db.create_bot_integration(
            bot_id=bot_id,
            integration_type="google_calendar",
            name="Google Calendar Cliente",
            config_data=config_data,
            enabled=enabled == "on"
        )
        
    # Save secrets if provided
    import re
    clean_secret = client_secret.strip()
    if clean_secret and not re.match(r"^\*+$", clean_secret):
        encrypted_secret = secure_store.encrypt_secret(clean_secret)
        await db.upsert_integration_secret(integration_id, "client_secret", encrypted_secret)
        
    clean_refresh = refresh_token.strip()
    if clean_refresh and not re.match(r"^\*+$", clean_refresh):
        encrypted_refresh = secure_store.encrypt_secret(clean_refresh)
        await db.upsert_integration_secret(integration_id, "refresh_token", encrypted_refresh)
        
    # Sync status also in bot_skills (for enabling calendar)
    await db.upsert_bot_skill(
        bot_id=bot_id,
        skill_type="google_calendar",
        enabled=enabled == "on",
        config_data={}
    )
    
    return RedirectResponse(f"/client/app?bot_id={bot_id}&tab=integrations&saved=1", status_code=302)

@router.post("/bots/{bot_id}/integrations/api")
async def client_api_save(
    request: Request,
    bot_id: int,
):
    session = _require_client_login(request)
    await _require_bot_editor(session, bot_id)
    
    form_data = await request.form()
    base_url = str(form_data.get("base_url", "")).strip()
    
    env_keys = form_data.getlist("env_key")
    env_vals = form_data.getlist("env_val")
        
    integration = await db.get_active_bot_integration(bot_id, "external_api")
    
    config_data = {
        "base_url": base_url,
    }
    
    if integration:
        integration_id = int(integration["id"])
        await db.update_bot_integration(
            bot_id=bot_id,
            integration_id=integration_id,
            integration_type="external_api",
            name=integration["name"],
            config_data=config_data,
            enabled=True
        )
    else:
        integration_id = await db.create_bot_integration(
            bot_id=bot_id,
            integration_type="external_api",
            name="API Cliente",
            config_data=config_data,
            enabled=True
        )
        
    # Handle dynamic secrets
    submitted_keys = []
    for k, v in zip(env_keys, env_vals):
        k_clean = str(k).strip()
        v_clean = str(v).strip()
        if not k_clean:
            continue
        submitted_keys.append(k_clean)
        
        # Only upsert if it's a new value (not the masked placeholder)
        if v_clean and not re.match(r"^\*+$", v_clean):
            encrypted_val = secure_store.encrypt_secret(v_clean)
            await db.upsert_integration_secret(integration_id, k_clean, encrypted_val)
            
    # Delete removed secrets
    existing_secrets = await db.list_integration_secrets(integration_id)
    for sec in existing_secrets:
        sec_name = sec["secret_name"]
        if sec_name not in submitted_keys:
            await db.delete_integration_secret(integration_id, sec_name)
        
    # Enable skill
    await db.upsert_bot_skill(
        bot_id=bot_id,
        skill_type="external_api",
        enabled=True,
        config_data={}
    )
    
    return RedirectResponse(f"/client/app?bot_id={bot_id}&tab=integrations&saved=1", status_code=302)

@router.post("/bots/{bot_id}/integrations/chatwoot")
async def client_chatwoot_save(
    request: Request,
    bot_id: int,
    enabled: str | None = Form(None),
    base_url: str = Form(""),
    account_id: str = Form(""),
    inbox_id: str = Form(""),
    api_token: str = Form(""),
):
    session = _require_client_login(request)
    await _require_bot_editor(session, bot_id)
    
    clean_base_url = base_url.strip()
    clean_account_id = account_id.strip()
    clean_inbox_id = inbox_id.strip()
    
    integration = await db.get_active_bot_integration(bot_id, "chatwoot")
    
    config_data = {
        "base_url": clean_base_url,
        "account_id": clean_account_id,
        "inbox_id": clean_inbox_id
    }
    
    is_enabled = (enabled == "on")
    
    if integration:
        integration_id = int(integration["id"])
        await db.update_bot_integration(
            bot_id=bot_id,
            integration_id=integration_id,
            integration_type="chatwoot",
            name=integration["name"],
            config_data=config_data,
            enabled=is_enabled
        )
    else:
        integration_id = await db.create_bot_integration(
            bot_id=bot_id,
            integration_type="chatwoot",
            name="Chatwoot Bandeja",
            config_data=config_data,
            enabled=is_enabled
        )
        
    clean_api_token = api_token.strip()
    if clean_api_token and not re.match(r"^\*+$", clean_api_token):
        encrypted_token = secure_store.encrypt_secret(clean_api_token)
        await db.upsert_integration_secret(integration_id, "api_token", encrypted_token)
        
@router.post("/bots/{bot_id}/integrations/routing")
async def client_routing_rules_save(
    request: Request,
    bot_id: int,
    enabled: str | None = Form(None),
    webhook_url: str = Form(""),
    phone_numbers: str = Form(""),
    save_history: str | None = Form(None),
    webhook_auth_token: str = Form(""),
):
    session = _require_client_login(request)
    await _require_bot_editor(session, bot_id)
    
    clean_webhook_url = webhook_url.strip()
    
    # Process and clean phone numbers list
    phones_list = []
    for p in phone_numbers.split(","):
        p_clean = p.replace("+", "").replace(" ", "").replace("-", "").strip()
        if p_clean:
            phones_list.append(p_clean)
            
    is_enabled = (enabled == "on")
    is_save_history = (save_history == "on")
    
    # Structure config data
    config_data = {
        "rules": [
            {
                "name": "Filtro de Omisión y Desvío",
                "filter_type": "phone_whitelist",
                "phone_numbers": phones_list,
                "action": "forward_and_bypass",
                "webhook_url": clean_webhook_url,
                "save_history": is_save_history
            }
        ]
    }
    
    # Find existing integration
    integration = await db.get_bot_integration_by_type(bot_id, "routing_rules")
    if integration:
        integration_id = int(integration["id"])
        await db.update_bot_integration(
            bot_id=bot_id,
            integration_id=integration_id,
            integration_type="routing_rules",
            name="Enrutamiento Externo",
            config_data=config_data,
            enabled=is_enabled
        )
    else:
        integration_id = await db.create_bot_integration(
            bot_id=bot_id,
            integration_type="routing_rules",
            name="Enrutamiento Externo",
            config_data=config_data,
            enabled=is_enabled
        )
        
    # Save auth token if submitted
    clean_token = webhook_auth_token.strip()
    if clean_token and not re.match(r"^\*+$", clean_token):
        encrypted_token = secure_store.encrypt_secret(clean_token)
        await db.upsert_integration_secret(integration_id, "webhook_auth_token", encrypted_token)
    elif not clean_token:
        # If token is empty, remove secret
        await db.delete_integration_secret(integration_id, "webhook_auth_token")
        
    return RedirectResponse(f"/client/app?bot_id={bot_id}&tab=integrations&saved=1", status_code=302)

@router.post("/bots/{bot_id}/skills/templates/install")
async def client_install_bot_skill_template(
    request: Request,
    bot_id: int,
    template_key: str = Form(...),
):
    session = _require_client_login(request)
    await _require_bot_editor(session, bot_id)
    from app.skill_templates import SKILL_TEMPLATES
    if template_key not in SKILL_TEMPLATES:
        raise HTTPException(status_code=404, detail="Plantilla no encontrada")
    
    tpl = SKILL_TEMPLATES[template_key]
    await db.upsert_bot_skill(
        bot_id=bot_id,
        skill_type=tpl["skill_type"],
        enabled=True,
        config_data=tpl["config"],
    )
    return RedirectResponse(f"/client/app?bot_id={bot_id}&tab=integrations&saved=1", status_code=302)

@router.post("/bots/{bot_id}/crm/{wa_id}/status")
async def client_crm_update_status(
    request: Request,
    bot_id: int,
    wa_id: str,
    status: str = Form(...),
):
    session = _require_client_login(request)
    await _require_bot_editor(session, bot_id)
    if status not in ("en_progreso", "calificado", "descalificado"):
        raise HTTPException(400, "Estado inválido")
    
    # Verify the lead belongs to this bot
    lead = await db.get_lead(wa_id, bot_id=bot_id)
    if not lead:
        raise HTTPException(404, "Lead no encontrado o no pertenece a este bot")
        
    await db.update_lead_status(
        wa_id,
        status,
        disqualify_reason="Movido manualmente desde panel cliente" if status == "descalificado" else None,
        bot_id=bot_id,
    )
    return RedirectResponse(f"/client/app?bot_id={bot_id}&tab=crm&status={status}", status_code=302)


@router.post("/bots/{bot_id}/whatsapp/templates")
async def client_whatsapp_templates_submit(
    request: Request,
    bot_id: int,
    name: str = Form(...),
    language: str = Form("es_MX"),
    category: str = Form("UTILITY"),
    body_text: str = Form(...),
    examples: list[str] = Form(None),
):
    session = _require_client_login(request)
    await _require_bot_editor(session, bot_id)
    if not body_text or not body_text.strip():
        return RedirectResponse(f"/client/app?bot_id={bot_id}&tab=templates&saved=err_El%20cuerpo%20de%20la%20plantilla%20no%20puede%20estar%20vac%C3%ADo.", status_code=302)
    try:
        await meta_provider.create_message_template(
            bot_id, name, language, category, body_text, examples=examples
        )
    except Exception as exc:
        log.error(f"Error al crear plantilla desde panel de cliente: {exc}")
        import urllib.parse
        safe_msg = urllib.parse.quote(str(exc))
        return RedirectResponse(f"/client/app?bot_id={bot_id}&tab=templates&saved=err_{safe_msg}", status_code=302)
    return RedirectResponse(f"/client/app?bot_id={bot_id}&tab=templates&saved=1", status_code=302)


# --- CONTACTS FILE PARSING & IMPORT HELPERS ---

def extract_headers_from_file(file_path: str) -> list[str]:
    _, ext = os.path.splitext(file_path.lower())
    try:
        if ext == ".csv":
            with open(file_path, "r", encoding="utf-8-sig") as f:
                reader = csv.reader(f)
                headers = next(reader, None)
                return [h.strip() for h in headers if h.strip()] if headers else []
        elif ext == ".xlsx":
            wb = openpyxl.load_workbook(file_path, read_only=True)
            sheet = wb.active
            for row in sheet.iter_rows(max_row=1, values_only=True):
                return [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
    except Exception as exc:
        log.error(f"Error reading headers from {file_path}: {exc}")
    return []


def parse_contacts_file(
    file_path: str,
    phone_col_idx: int,
    name_col_idx: int | None = None,
    business_col_idx: int | None = None,
    tags_col_idx: int | None = None,
) -> list[dict]:
    contacts = []
    _, ext = os.path.splitext(file_path.lower())
    try:
        if ext == ".csv":
            with open(file_path, "r", encoding="utf-8-sig") as f:
                reader = csv.reader(f)
                next(reader, None)  # Skip header
                for row in reader:
                    if not row or len(row) <= phone_col_idx:
                        continue
                    phone = row[phone_col_idx].strip()
                    if not phone:
                        continue
                    name = row[name_col_idx].strip() if (name_col_idx is not None and len(row) > name_col_idx) else None
                    business = row[business_col_idx].strip() if (business_col_idx is not None and len(row) > business_col_idx) else None
                    tags = row[tags_col_idx].strip() if (tags_col_idx is not None and len(row) > tags_col_idx) else None
                    contacts.append({
                        "wa_id": phone,
                        "name": name,
                        "business": business,
                        "tags": tags,
                    })
        elif ext == ".xlsx":
            wb = openpyxl.load_workbook(file_path, data_only=True)
            sheet = wb.active
            first = True
            for row in sheet.iter_rows(values_only=True):
                if first:
                    first = False
                    continue
                if not row or len(row) <= phone_col_idx:
                    continue
                phone = str(row[phone_col_idx] or "").strip()
                if not phone:
                    continue
                name = str(row[name_col_idx] or "").strip() if (name_col_idx is not None and len(row) > name_col_idx) else None
                business = str(row[business_col_idx] or "").strip() if (business_col_idx is not None and len(row) > business_col_idx) else None
                tags = str(row[tags_col_idx] or "").strip() if (tags_col_idx is not None and len(row) > tags_col_idx) else None
                contacts.append({
                    "wa_id": phone,
                    "name": name,
                    "business": business,
                    "tags": tags,
                })
    except Exception as exc:
        log.error(f"Error parsing contacts file {file_path}: {exc}")
    return contacts


# --- CONTACTS & CAMPAIGNS ENDPOINTS ---

@router.post("/bots/{bot_id}/contacts/upload")
async def client_contacts_upload(
    request: Request,
    bot_id: int,
    file: UploadFile = File(...),
):
    session = _require_client_login(request)
    await _require_bot_editor(session, bot_id)
    
    _, ext = os.path.splitext(file.filename.lower())
    if ext not in (".csv", ".xlsx"):
        return RedirectResponse(
            f"/client/app?bot_id={bot_id}&tab=contacts&saved=err_ext", status_code=302
        )
        
    # Ensure tmp directory exists inside workspace
    tmp_dir = os.path.join(os.getcwd(), "tmp", "uploads")
    os.makedirs(tmp_dir, exist_ok=True)
    
    # Save the file with a unique name
    file_token = f"{uuid.uuid4()}{ext}"
    file_path = os.path.join(tmp_dir, file_token)
    try:
        await file.seek(0)
        with open(file_path, "wb") as f:
            content = await file.read()
            if len(content) > config.CONTACTS_UPLOAD_MAX_BYTES:
                return RedirectResponse(
                    f"/client/app?bot_id={bot_id}&tab=contacts&saved=err_file_too_large",
                    status_code=302,
                )
            f.write(content)
    except Exception as exc:
        log.error(f"Error saving uploaded contacts file: {exc}")
        return RedirectResponse(
            f"/client/app?bot_id={bot_id}&tab=contacts&saved=err", status_code=302
        )
        
    return RedirectResponse(
        f"/client/app?bot_id={bot_id}&tab=contacts&map_file={file_token}", status_code=302
    )


@router.post("/bots/{bot_id}/contacts/import")
async def client_contacts_confirm_import(
    request: Request,
    bot_id: int,
    file_token: str = Form(...),
    phone_col: int = Form(...),
    name_col: int = Form(-1),
    business_col: int = Form(-1),
    tags_col: int = Form(-1),
):
    session = _require_client_login(request)
    await _require_bot_editor(session, bot_id)
    
    tmp_dir = os.path.join(os.getcwd(), "tmp", "uploads")
    file_path = os.path.join(tmp_dir, file_token)
    if not os.path.exists(file_path):
        return RedirectResponse(
            f"/client/app?bot_id={bot_id}&tab=contacts&saved=err_file_expired", status_code=302
        )
        
    n_col = None if name_col < 0 else name_col
    b_col = None if business_col < 0 else business_col
    t_col = None if tags_col < 0 else tags_col
    
    parsed = parse_contacts_file(
        file_path,
        phone_col_idx=phone_col,
        name_col_idx=n_col,
        business_col_idx=b_col,
        tags_col_idx=t_col,
    )
    
    success_count = 0
    for c in parsed:
        wa_id = "".join(filter(str.isdigit, c["wa_id"]))
        if wa_id:
            try:
                await db.upsert_contact(
                    bot_id,
                    wa_id=wa_id,
                    name=c.get("name"),
                    business=c.get("business"),
                    tags=c.get("tags"),
                )
                success_count += 1
            except Exception as exc:
                log.error(f"Error importing contact {wa_id}: {exc}")
                
    # Delete temporary file
    try:
        os.remove(file_path)
    except Exception as exc:
        log.error(f"Error removing temporary file {file_path}: {exc}")
        
    return RedirectResponse(
        f"/client/app?bot_id={bot_id}&tab=contacts&imported={success_count}", status_code=302
    )


@router.post("/bots/{bot_id}/contacts/create-manual")
async def client_contacts_create_manual(
    request: Request,
    bot_id: int,
    name: str = Form(...),
    wa_id: str = Form(...),
    business: str = Form(None),
    tags: str = Form(None),
):
    session = _require_client_login(request)
    await _require_bot_editor(session, bot_id)
    
    clean_wa = "".join(filter(str.isdigit, wa_id))
    if not clean_wa:
        return RedirectResponse(
            f"/client/app?bot_id={bot_id}&tab=contacts&saved=err_num", status_code=302
        )
        
    try:
        await db.upsert_contact(
            bot_id,
            wa_id=clean_wa,
            name=name.strip() if name else None,
            business=business.strip() if business else None,
            tags=tags.strip() if tags else None,
        )
    except Exception as exc:
        log.error(f"Error creating contact manually: {exc}")
        return RedirectResponse(
            f"/client/app?bot_id={bot_id}&tab=contacts&saved=err", status_code=302
        )
        
    return RedirectResponse(
        f"/client/app?bot_id={bot_id}&tab=contacts&saved=1", status_code=302
    )


@router.post("/bots/{bot_id}/contacts/delete")
async def client_contacts_delete(
    request: Request,
    bot_id: int,
    selected_contacts: list[str] = Form(...),
):
    session = _require_client_login(request)
    await _require_bot_editor(session, bot_id)
    
    try:
        await db.delete_contacts(bot_id, selected_contacts)
    except Exception as exc:
        log.error(f"Error deleting contacts: {exc}")
        return RedirectResponse(
            f"/client/app?bot_id={bot_id}&tab=contacts&saved=err", status_code=302
        )
        
    return RedirectResponse(
        f"/client/app?bot_id={bot_id}&tab=contacts&deleted={len(selected_contacts)}", status_code=302
    )


async def process_broadcast_queue(broadcast_id: int, bot_id: int) -> None:
    """Procesa en segundo plano los envíos individuales de una campaña masiva."""
    log.info(f"Iniciando procesamiento de campaña masiva {broadcast_id} para bot {bot_id}...")
    try:
        # Actualizar campaña a estado 'running'
        await db.update_broadcast_status(broadcast_id, "running")
        
        # Obtener los detalles de la campaña
        broadcast = await db.get_broadcast(broadcast_id, bot_id)
        if not broadcast:
            log.error(f"Campaña {broadcast_id} no encontrada en base de datos.")
            return
            
        template_name = broadcast["template_name"]
        lang_code = broadcast["language_code"]
        mappings = json.loads(broadcast["variable_mappings"])
        
        while True:
            # Obtener lote de 50 destinatarios pendientes
            recipients = await db.get_pending_broadcast_recipients(broadcast_id, limit=50)
            if not recipients:
                break
                
            for r in recipients:
                recipient_id = int(r["id"])
                wa_id = r["wa_id"]
                
                # Resolver variables para este destinatario específico
                resolved_params = []
                for m in mappings:
                    m_type = m.get("type")
                    m_val = m.get("value")
                    
                    if m_type == "fixed":
                        resolved_params.append(m_val or "")
                    elif m_type == "name":
                        resolved_params.append(r.get("contact_name") or "")
                    elif m_type == "business":
                        resolved_params.append(r.get("contact_business") or "")
                    elif m_type == "wa_id":
                        resolved_params.append(wa_id)
                    else:
                        resolved_params.append("")
                        
                # Llamar API de Meta
                try:
                    await meta_provider.send_template_message(
                        bot_id=bot_id,
                        to_wa_id=wa_id,
                        template_name=template_name,
                        language_code=lang_code,
                        parameters=resolved_params,
                    )
                    await db.update_broadcast_recipient_status(recipient_id, "sent")
                except Exception as exc:
                    log.error(f"Error al enviar mensaje de campaña a {wa_id}: {exc}")
                    await db.update_broadcast_recipient_status(recipient_id, "failed", error_message=str(exc))
                    
                await asyncio.sleep(0.1)
                
        # Finalizar campaña
        await db.update_broadcast_status(broadcast_id, "completed")
        log.info(f"Procesamiento de campaña {broadcast_id} completado con éxito.")
    except Exception as exc:
        log.error(f"Error en worker de campaña masiva {broadcast_id}: {exc}")
        await db.update_broadcast_status(broadcast_id, "failed")


@router.post("/bots/{bot_id}/campaigns/create")
async def client_campaigns_create(
    request: Request,
    bot_id: int,
    background_tasks: BackgroundTasks,
    campaign_name: str = Form(...),
    template_name: str = Form(...),
    language_code: str = Form("es_MX"),
    recipients_option: str = Form("selected"), # "selected" o "all"
    selected_wa_ids: str = Form(""),
    vars_count: int = Form(0),
    confirm_send: str = Form(""),
):
    session = _require_client_login(request)
    await _require_bot_editor(session, bot_id)
    
    recipients_list = []
    try:
        if recipients_option == "all":
            contacts = await db.list_contacts(bot_id, limit=10000)
            recipients_list = [
                {"wa_id": c["wa_id"], "name": c.get("name"), "business": c.get("business")}
                for c in contacts
            ]
        else:
            clean_ids = [w.strip() for w in selected_wa_ids.split(",") if w.strip()]
            contacts = await db.list_contacts(bot_id, limit=10000)
            contacts_map = {c["wa_id"]: c for c in contacts}
            for wa in clean_ids:
                c = contacts_map.get(wa) or {}
                recipients_list.append({
                    "wa_id": wa,
                    "name": c.get("name"),
                    "business": c.get("business"),
                })
    except Exception as exc:
        log.error(f"Error resolving campaign recipients: {exc}")
        return RedirectResponse(
            f"/client/app?bot_id={bot_id}&tab=campaigns&saved=err", status_code=302
        )
        
    if not recipients_list:
        return RedirectResponse(
            f"/client/app?bot_id={bot_id}&tab=campaigns&saved=err_no_recipients", status_code=302
        )
    if confirm_send.strip().upper() != "CONFIRMAR":
        return RedirectResponse(
            f"/client/app?bot_id={bot_id}&tab=campaigns&saved=err_confirm", status_code=302
        )
    if len(recipients_list) > config.CAMPAIGN_MAX_RECIPIENTS:
        return RedirectResponse(
            f"/client/app?bot_id={bot_id}&tab=campaigns&saved=err_campaign_limit", status_code=302
        )
        
    variable_mappings = []
    form_data = await request.form()
    for i in range(1, vars_count + 1):
        map_type = form_data.get(f"var_map_type_{i}") or "fixed"
        map_value = form_data.get(f"var_map_value_{i}") or ""
        variable_mappings.append({
            "var_idx": i,
            "type": map_type,
            "value": map_value if map_type == "fixed" else "",
        })
        
    try:
        broadcast_id = await db.create_broadcast(
            bot_id=bot_id,
            name=campaign_name.strip(),
            template_name=template_name,
            language_code=language_code,
            variable_mappings=variable_mappings,
            recipients=recipients_list,
        )
    except Exception as exc:
        log.error(f"Error creating broadcast record: {exc}")
        return RedirectResponse(
            f"/client/app?bot_id={bot_id}&tab=campaigns&saved=err", status_code=302
        )
        
    background_tasks.add_task(process_broadcast_queue, broadcast_id, bot_id)
    
    return RedirectResponse(
        f"/client/app?bot_id={bot_id}&tab=campaigns&saved=1", status_code=302
    )

def _render_skill_templates(bot_id: int, session: dict) -> str:
    from app.skill_templates import SKILL_TEMPLATES
    html_cards = []
    for key, tpl in SKILL_TEMPLATES.items():
        btn = (
            f'<button class="btn primary-btn" type="submit">Instalar Plantilla</button>'
            if session.get("role") != "client_viewer" else
            '<span class="badge">Solo lectura</span>'
        )
        html_cards.append(
            f"""
            <div class="card" style="border: 1px dashed var(--line-strong); display:flex; flex-direction:column; justify-content:space-between; padding: 20px; box-shadow: none;">
              <div>
                <h3 style="color:var(--primary); margin: 0 0 6px 0; font-family:Outfit; font-size:17px; font-weight:700;">{html.escape(tpl["name"])}</h3>
                <p style="font-size:12.5px; color:var(--muted); line-height:1.45; margin-bottom:14px;">{tpl["description"]}</p>
              </div>
              <form method="post" action="/client/bots/{bot_id}/skills/templates/install" style="margin-top:auto; padding:0;">
                <input type="hidden" name="template_key" value="{html.escape(key)}">
                <div class="actions" style="margin:0; padding:0;">{btn}</div>
              </form>
            </div>
            """
        )
    return "\n".join(html_cards)

import io
import csv
import logging

log = logging.getLogger("file-parser")

# Dynamic imports inside functions to fail gracefully if needed
def parse_file(file_bytes: bytes, filename: str) -> str:
    """Parses text content from file bytes based on filename extension.
    Supported extensions: .pdf, .docx, .xlsx, .csv, .md, .txt
    """
    ext = filename.lower().split('.')[-1]
    
    if ext == "pdf":
        try:
            import pypdf
            reader = pypdf.PdfReader(io.BytesIO(file_bytes))
            text_parts = []
            for page in reader.pages:
                text = page.extract_text()
                if text:
                    text_parts.append(text)
            return "\n".join(text_parts)
        except Exception as e:
            log.error(f"Error parsing PDF file: {e}")
            raise ValueError(f"No se pudo leer el archivo PDF: {e}")
            
    elif ext in ("docx", "doc"):
        try:
            import docx
            doc = docx.Document(io.BytesIO(file_bytes))
            return "\n".join([p.text for p in doc.paragraphs])
        except Exception as e:
            log.error(f"Error parsing DOCX file: {e}")
            raise ValueError(f"No se pudo leer el archivo Word: {e}")
            
    elif ext == "xlsx":
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
            sheet_texts = []
            for sheet in wb.sheetnames:
                sheet_texts.append(f"--- Hoja: {sheet} ---")
                ws = wb[sheet]
                for row in ws.iter_rows(values_only=True):
                    if any(val is not None for val in row):
                        sheet_texts.append(", ".join([str(val) if val is not None else "" for val in row]))
            return "\n".join(sheet_texts)
        except Exception as e:
            log.error(f"Error parsing XLSX file: {e}")
            raise ValueError(f"No se pudo leer el archivo Excel: {e}")
            
    elif ext == "csv":
        try:
            decoded = file_bytes.decode('utf-8', errors='ignore')
            reader = csv.reader(io.StringIO(decoded))
            rows = []
            for row in reader:
                if any(row):
                    rows.append(", ".join(row))
            return "\n".join(rows)
        except Exception as e:
            log.error(f"Error parsing CSV file: {e}")
            raise ValueError(f"No se pudo leer el archivo CSV: {e}")
            
    elif ext in ("md", "txt"):
        try:
            return file_bytes.decode('utf-8', errors='ignore')
        except Exception as e:
            log.error(f"Error parsing plain text file: {e}")
            raise ValueError(f"No se pudo leer el archivo de texto: {e}")
            
    else:
        raise ValueError(f"Formato de archivo no soportado: .{ext}")
